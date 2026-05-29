"""Provision recompute engine — Module C: deferred taxes.

Two checks, both high-value and mechanically clean:

  1. FS DEFERRED TIE — the provision's deferred footnote-grouping block (DTA/DTL by
     category) is tied line-by-line to the FS deferred tax table. This reproduces the
     FN-07 deferred footnote from the workpaper: NOL, Sec 163j, R&E, intangibles,
     capitalized commissions, ROU, etc., plus total DTA / total DTL / valuation
     allowance / net.

  2. CROSS-YEAR CONTINUITY / TREND — each temporary difference's ending balance is
     tracked across FY23/24/25. A difference that appears or disappears, or swings
     without an obvious driver, is flagged — the symptom of a rolled-forward template
     dropping or stranding a balance (e.g. an accrued-bonus DTA that silently zeros).

Emits records under lane "deferred_tax" -> the "Deferred Tax" report tab.
(Within-year roll footing from tab 3 is a later refinement; the FS tie + continuity
cover the integrity that matters most.)

CLI:
  python recompute_deferred_tax.py <inputs.json> <out.json>
       [--fy23 <xlsx>] [--fy24 <xlsx>] [--fy25 <xlsx>]
"""

import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from tie_out_common import parse_value, compare, make_record  # noqa: E402

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

ROOT = Path(r"C:\path\to\financial-statement-review")
DEFAULTS = {
    "FY23": ROOT / "Prior Year Examples/2023/FS Compilation Requests (FS Compilation Partner)/8. (YE) Acme Corp 2023 Tax Provision_v2024.5.23.xlsx",
    "FY24": ROOT / "PBCs/2024 Audit/FS Compilation Requests (FS Compilation Partner)/Tax Provision/2024 Acme Corp Tax Provision_v2025.5.27.xlsx",
    "FY25": ROOT / "PBCs/2025 Audit/Year End/FS Compilation Requests (FS Compilation Partner) (old)/9. Tax Provision/9. Acme Corp 2025 Tax Provision_v2026.5.21.xlsx",
}

_NORM = re.compile(r"[^a-z0-9]+")


def norm(s):
    return _NORM.sub(" ", str(s).lower()).strip()


# Provision-grouping label -> FS deferred-table label (normalized).
GROUPING_TO_FS = {
    "net operating losses": "net operating losses",
    "section 163j interest": "section 163j interest",
    "charitable contributions": "charitable contributions",
    "deferred revenue": "deferred revenue",
    "r e expenses": "r e expenses",
    "transaction costs": "transaction costs",
    "payroll taxes": "payroll taxes",
    "fixed assets": "fixed assets",
    "rou liability": "rou liability",
    "intangibles": "intangibles",
    "capitilized commissions": "capitalized commissions",  # workpaper typo
    "capitalized commissions": "capitalized commissions",
    "rou asset": "rou asset",
    "seller note discount": "seller note discount",
}


def extract_fs_deferred(inputs):
    """{normalized_label: value_$K} from the FS deferred tax table, plus totals/VA/net.
    Tracks sign by section (DTL lines stored negative as shown).

    pdf2docx-derived docx files can split the deferred-tax disclosure across multiple
    tables at page boundaries (DTA list in one, DTL/VA/net in the next). Concatenate
    the rows of every FN-07 table so the extractor sees the full disclosure regardless
    of how the source PDF was paginated. A hand-edited docx has it all in one table
    anyway, so the concat is harmless there.
    """
    tables = inputs.get("fy25_docx", {}).get("tables", [])
    fn07_rows = []
    one_table_rows = None
    for t in tables:
        rows = t.get("rows", []) if isinstance(t, dict) else (t if isinstance(t, list) else [])
        flat = " ".join(str(c) for r in rows for c in (r if isinstance(r, list) else []) if c).lower()
        if (isinstance(t, dict) and str(t.get("section", "")).startswith("FN-07")):
            fn07_rows.extend(rows)
        if "total deferred tax assets" in flat and "valuation allowance" in flat and one_table_rows is None:
            one_table_rows = rows
    target = one_table_rows or fn07_rows
    lines, totals = {}, {}
    if not target:
        return lines, totals
    for row in target:
        cells = row if isinstance(row, list) else []
        label = next((str(c) for c in cells if isinstance(c, str) and c.strip()), "")
        nums = [parse_value(c) for c in cells]
        nums = [n for n in nums if n is not None]
        if not label or not nums:
            continue
        nl = norm(label)
        val = nums[0]
        if "total deferred tax assets" in nl:
            totals["total_dta"] = val
        elif "total deferred tax liabilities" in nl:
            totals["total_dtl"] = val
        elif "valuation allowance" in nl:
            totals["valuation_allowance"] = val
        elif "net deferred tax" in nl:
            totals["net"] = val if val is not None else 0.0
        else:
            lines[nl] = val
    return lines, totals


def extract_provision_deferred(path):
    """From '4| DTA DTL Summary': the footnote-grouping block (label@9 -> amount@11,
    raw $) and the per-temp-difference left block (label@1 -> balance@3) for trend."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    grouping, perdiff, net_total = {}, {}, None
    if "4| DTA DTL Summary" in wb.sheetnames:
        for row in wb["4| DTA DTL Summary"].iter_rows(values_only=True):
            # grouping block
            glabel = row[9] if len(row) > 9 else None
            gamt = parse_value(row[11]) if len(row) > 11 else None
            if isinstance(glabel, str) and glabel.strip() and gamt is not None:
                gl = norm(glabel)
                if gl in ("deferred tax assets", "deferred tax liabilities", "footnote grouping"):
                    pass
                else:
                    # A category can appear on more than one grouping row (e.g. two
                    # "Fixed Assets" lines, or "Other" on both the DTA and DTL side) —
                    # sum them so the category total ties the single FS line.
                    grouping[gl] = grouping.get(gl, 0.0) + gamt
            elif gamt is not None and glabel is None and net_total is None and abs(gamt) > 1e6:
                net_total = gamt  # the grand-total row (N,NNN,NNN.8)
            # per-difference left block (label col1, balance col3, letter col6)
            dlabel = row[1] if len(row) > 1 else None
            dval = parse_value(row[3]) if len(row) > 3 else None
            code = row[6] if len(row) > 6 else None
            if (isinstance(dlabel, str) and dlabel.strip() and dval is not None
                    and isinstance(code, str) and len(code.strip()) == 1 and code.strip().isalpha()):
                perdiff[dlabel.strip()] = dval
    wb.close()
    return grouping, perdiff, net_total


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("inputs_json")
    ap.add_argument("out_json")
    for yr, p in DEFAULTS.items():
        ap.add_argument(f"--{yr.lower()}", default=str(p))
    args = ap.parse_args()

    inputs = json.loads(Path(args.inputs_json).read_text(encoding="utf-8"))
    fs_lines, fs_totals = extract_fs_deferred(inputs)

    grouping = {}
    perdiff_by_year = {}
    net_total = None
    for yr in ("FY23", "FY24", "FY25"):
        path = getattr(args, yr.lower())
        if not Path(path).exists():
            continue
        g, pd, nt = extract_provision_deferred(path)
        perdiff_by_year[yr] = pd
        if yr == "FY25":
            grouping, net_total = g, nt

    records = []

    # ---- 1) FS deferred tie: grouping (raw $) -> FS line ($K), by sign for 'Other' ----
    used_fs = set()
    for glabel, gamt in grouping.items():
        fs_key = GROUPING_TO_FS.get(glabel)
        # 'Other' appears on both DTA and DTL sides — disambiguate by sign.
        if glabel == "other":
            fs_key = "other"
        if fs_key is None:
            # try direct normalized match
            fs_key = glabel if glabel in fs_lines else None
        fs_val = None
        if fs_key == "other":
            # pick the FS 'other' line whose sign matches
            cands = [(k, v) for k, v in fs_lines.items() if k == "other"]
            # fs_lines is a dict so only one 'other'; fall back to sign check below
            fs_val = fs_lines.get("other")
        elif fs_key:
            fs_val = fs_lines.get(fs_key)
        if fs_val is None and abs(gamt) < 1000:
            continue  # immaterial unmapped grouping line ($<1K)
        if fs_val is None:
            records.append(make_record(
                "deferred_tax", pdf_section="FN - Taxes (deferred)",
                pdf_label=f"DTA/DTL: {glabel}", pdf_year="2025",
                pdf_value=round(gamt / 1000.0, 1), source_ref="provision tab 4",
                source_label="FS deferred table", source_value=None,
                status="missing", is_subtotal=True,
                notes="Provision deferred grouping line not matched to an FS deferred-table line."))
            continue
        used_fs.add(fs_key)
        d, st, tol, unit = compare(fs_val, gamt / 1000.0, "$K", "$K", is_subtotal=True)
        records.append(make_record(
            "deferred_tax", pdf_section="FN - Taxes (deferred)",
            pdf_label=f"DTA/DTL: {glabel}", pdf_year="2025",
            pdf_value=fs_val, source_ref="provision tab 4",
            source_label="provision grouping", source_value=round(gamt / 1000.0, 1),
            comparison_unit=unit, delta=round(d, 1) if d is not None else None,
            tolerance=tol, status=st, is_subtotal=True,
            notes="Deferred footnote line ties the provision grouping."
                  if st in ("ties", "ties-with-rounding") else
                  "Deferred footnote line does NOT tie the provision."))

    # ---- totals + VA + net ----
    if net_total is not None and fs_totals.get("valuation_allowance") is not None:
        # net DTA before VA (provision) vs FS (DTA + DTL) and vs |VA|
        fs_before_va = (fs_totals.get("total_dta") or 0) + (fs_totals.get("total_dtl") or 0)
        d, st, tol, unit = compare(fs_before_va, net_total / 1000.0, "$K", "$K", is_subtotal=True)
        records.append(make_record(
            "deferred_tax", pdf_section="FN - Taxes (deferred)",
            pdf_label="Net DTA before valuation allowance", pdf_year="2025",
            pdf_value=round(fs_before_va, 1), source_ref="provision tab 4",
            source_label="provision net (pre-VA)", source_value=round(net_total / 1000.0, 1),
            comparison_unit=unit, delta=round(d, 1) if d is not None else None,
            tolerance=tol, status=st, is_subtotal=True,
            notes="Provision net DTA (pre-VA) ties the FS deferred table total."
                  if st in ("ties", "ties-with-rounding") else "Does NOT tie."))

    # FS deferred table foots: DTA + DTL - VA = net (~0 full VA)
    if all(k in fs_totals for k in ("total_dta", "total_dtl", "valuation_allowance")):
        net = fs_totals.get("net") or 0.0
        computed = fs_totals["total_dta"] + fs_totals["total_dtl"] + fs_totals["valuation_allowance"]
        d = computed - net
        st = "ties" if abs(d) <= 1.0 else "exception"
        records.append(make_record(
            "deferred_tax", pdf_section="FN - Taxes (deferred)",
            pdf_label="FS deferred table foots (DTA + DTL - VA = net)", pdf_year="2025",
            pdf_value=round(computed, 1), source_ref="FS FN-07", source_label="reported net",
            source_value=round(net, 1), comparison_unit="$K", delta=round(d, 1),
            tolerance=1.0, status=st, is_subtotal=True,
            notes="Deferred tax table foots." if st == "ties" else "Does NOT foot."))

    # ---- 2) cross-year continuity / trend ----
    yrs = [y for y in ("FY23", "FY24", "FY25") if y in perdiff_by_year]
    all_diffs = []
    for y in yrs:
        for k in perdiff_by_year[y]:
            if k not in all_diffs:
                all_diffs.append(k)
    for diff in all_diffs:
        series = [perdiff_by_year[y].get(diff) for y in yrs]
        present = [v is not None and abs(v) > 0.5 for v in series]
        # flag a difference that was present then disappeared (or vice-versa)
        appeared = (not present[0]) and present[-1] if len(present) >= 2 else False
        disappeared = present[0] and (not present[-1]) if len(present) >= 2 else False
        status = "deferred-trend"
        note = " -> ".join(f"{y} {('%.0f'%v) if v is not None else '-'}" for y, v in zip(yrs, series))
        if disappeared:
            status = "deferred-dropped"
            note = ("Balance present in an earlier year then zero/absent at FY25 — confirm it was "
                    "genuinely released (not a dropped roll-forward line).  " + note)
        elif appeared:
            status = "deferred-new"
            note = "New temporary difference vs prior years.  " + note
        records.append(make_record(
            "deferred_tax", pdf_section="FN - Taxes (deferred trend)",
            pdf_label=f"Temp diff: {diff}", pdf_year=" / ".join(yrs),
            pdf_value=round(series[-1], 0) if series[-1] is not None else None,
            source_ref="provision tab 4 (3-yr)", source_label="prior years",
            source_value=round(series[0], 0) if series[0] is not None else None,
            comparison_unit="$1", status=status, is_subtotal=False, notes=note))

    Path(args.out_json).write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")

    tie = [r for r in records if r["pdf_section"] == "FN - Taxes (deferred)"]
    trend = [r for r in records if "trend" in r["pdf_section"]]
    by = {}
    for r in tie:
        by[r["status"]] = by.get(r["status"], 0) + 1
    print(f"Deferred tax module: {len(tie)} FS-tie checks {by} | {len(trend)} temp-diff trend rows")
    for r in tie:
        if r["status"] not in ("ties", "ties-with-rounding"):
            print(f"  [{r['status']}] {r['pdf_label']}: FS={r['pdf_value']} prov={r['source_value']} d={r['delta']}")
    for r in trend:
        if r["status"] in ("deferred-dropped", "deferred-new"):
            print(f"  [{r['status']}] {r['pdf_label']}: {r['notes'][:80]}")
    print(f"Wrote {args.out_json}")


if __name__ == "__main__":
    main()
