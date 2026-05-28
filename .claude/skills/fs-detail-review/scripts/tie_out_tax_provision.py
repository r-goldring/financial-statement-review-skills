"""Phase 2 — Tax provision module (FN-07).

Ties the income-tax footnote back to the provision workpaper the tax preparer built
(the version the compiler folded into the bridge/FS), and re-performs the checks an
auditor runs on a provision:

  1. Statutory recompute — pretax book income x federal rate = "tax at statutory rate".
  2. Rate-rec footing — the statutory-to-effective reconciliation foots to the total
     provision (the workpaper's own "Reconciliation - s/b zero" check).
  3. Total provision <-> FS income tax expense.
  4. Book pretax reconciliation — the pretax the provision is built on must equal the
     FS pretax. (This is the $X,XXX.XXK class of error: someone updates the income
     statement but not the provision input, or the footnote pretax drifts from the
     face. Caught here at the source.)
  5. Deferred summary — net DTA/DTL before valuation allowance ties the FS deferred
     table, and the FS table foots (DTA + DTL - VA = net, ~0 for a full VA).

Emits records under lane "tax_provision" -> the "Tax Provision" report tab.

CLI:
  python tie_out_tax_provision.py <inputs.json> <pbc-index.json> <out.json>
"""

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from tie_out_common import parse_value, compare, make_record  # noqa: E402

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────────────
# FS-side extraction (from inputs.json fy25_docx.tables) — values in $K
# ─────────────────────────────────────────────────────────────────────────────
def _table_rows(t):
    return t if isinstance(t, list) else t.get("rows", [])


def fs_find_row_value(tables, label_substrs, col=0):
    """Find the first docx-table row whose first cell matches all label_substrs,
    return the `col`-th numeric value on that row (0 = first number)."""
    for t in tables:
        for row in _table_rows(t):
            cells = row if isinstance(row, list) else []
            label = next((str(c) for c in cells if isinstance(c, str) and c.strip()), "")
            ll = label.lower()
            if all(s.lower() in ll for s in label_substrs):
                nums = [parse_value(c) for c in cells]
                nums = [n for n in nums if n is not None]
                if len(nums) > col:
                    return nums[col]
    return None


def extract_fs_tax(inputs):
    tables = inputs.get("fy25_docx", {}).get("tables", [])
    return {
        "pretax": fs_find_row_value(tables, ["loss before income tax"]),
        "tax_expense": fs_find_row_value(tables, ["income tax expense"]),
        "total_current": fs_find_row_value(tables, ["total current"]),
        "total_deferred": fs_find_row_value(tables, ["total deferred"]),
        "total_dta": fs_find_row_value(tables, ["total deferred tax assets"]),
        "total_dtl": fs_find_row_value(tables, ["total deferred tax liabilities"]),
        "valuation_allowance": fs_find_row_value(tables, ["valuation allowance"]),
        "net_dta": fs_find_row_value(tables, ["net deferred tax asset"]),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Provision-workbook extraction (raw $) — '6| Rate Rec.' and '4| DTA DTL Summary'
# ─────────────────────────────────────────────────────────────────────────────
def _ws_rows(wb, name):
    ws = wb[name]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _row_first_number(rows, label_substr, skip_rate=True):
    """First numeric on the row whose label (cols 0-2) contains label_substr.
    If skip_rate, ignore pure-fraction cells (|v|<1) so a rate column isn't picked."""
    for row in rows:
        label = next((str(c) for c in row[:3] if isinstance(c, str) and c.strip()), "")
        if label_substr.lower() in label.lower():
            for c in row:
                v = parse_value(c)
                if v is None:
                    continue
                if skip_rate and abs(v) < 1:
                    continue
                return v
    return None


def extract_provision(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rr = _ws_rows(wb, "6| Rate Rec.") if "6| Rate Rec." in wb.sheetnames else []
    dta = _ws_rows(wb, "4| DTA DTL Summary") if "4| DTA DTL Summary" in wb.sheetnames else []

    # Federal rate: the fractional value on the "Federal Rate" row (|v|<1).
    fed_rate = None
    for row in rr:
        label = " ".join(str(c) for c in row if isinstance(c, str))
        if "federal rate" in label.lower():
            for c in row:
                v = parse_value(c)
                if v is not None and 0 < abs(v) < 1:
                    fed_rate = v
                    break
        if fed_rate:
            break

    # Net DTA before VA: sum the per-difference "Total Balance" column (col index 3)
    # across the temporary-difference DETAIL rows only. Each detail row carries a
    # single-letter footnote-grouping code in col index 6 ('A','B','C',...); the
    # total/subtotal rows do not — so the letter gate excludes them (otherwise the
    # total row would be double-counted with the details).
    net_dta = 0.0
    counted = 0
    for row in dta:
        label = row[1] if len(row) > 1 else None
        val = parse_value(row[3]) if len(row) > 3 else None
        code = row[6] if len(row) > 6 else None
        is_detail = isinstance(code, str) and len(code.strip()) == 1 and code.strip().isalpha()
        if isinstance(label, str) and label.strip() and val is not None and is_detail:
            net_dta += val
            counted += 1

    wb.close()
    return {
        "pretax": _row_first_number(rr, "Pre-tax Book Income"),
        "federal_rate": fed_rate,
        "statutory_tax": _row_first_number(rr, "Tax at Statutory Rate"),
        "total_provision": _row_first_number(rr, "Per Account Rollforward"),
        "recon_check": _row_first_number(rr, "Reconciliation - s/b zero", skip_rate=False),
        "net_dta_before_va": net_dta if counted else None,
    }


# ─────────────────────────────────────────────────────────────────────────────
def add(records, label, fs_val_K, prov_val_K, ref, note_ok, note_bad,
        is_subtotal=True, exact=False):
    """Emit one tie record comparing an FS $K value to a provision $K value."""
    if fs_val_K is None or prov_val_K is None:
        records.append(make_record(
            "tax_provision", pdf_section="FN - Taxes", pdf_label=label,
            pdf_year="2025", pdf_value=fs_val_K, source_ref=ref,
            source_label="provision workpaper", source_value=prov_val_K,
            status="missing", is_subtotal=is_subtotal,
            notes="Could not locate one side of the comparison.",
        ))
        return None
    kind = "internal" if exact else None
    d, st, tol, unit = compare(fs_val_K, prov_val_K, "$K", "$K",
                               is_subtotal=is_subtotal, kind=kind)
    records.append(make_record(
        "tax_provision", pdf_section="FN - Taxes", pdf_label=label, pdf_year="2025",
        pdf_value=round(fs_val_K, 1), source_ref=ref, source_label="provision workpaper",
        source_value=round(prov_val_K, 1), comparison_unit=unit,
        delta=round(d, 1) if d is not None else None, tolerance=tol, status=st,
        is_subtotal=is_subtotal,
        notes=note_ok if st in ("ties", "ties-with-rounding") else note_bad,
    ))
    return st


def find_provision_wb(pbc_index):
    cands = [r for r in pbc_index if r.get("category") == "tax" and r.get("fs_relevant")
             and r.get("ext") in (".xlsx", ".xlsm")
             and "tax provision" in r["rel"].lower() and "9. tax provision" in r["rel"].lower()]
    if not cands:
        # fallback: any tax workbook whose filename has "Tax Provision"
        cands = [r for r in pbc_index if r.get("category") == "tax"
                 and r.get("ext") in (".xlsx", ".xlsm")
                 and "tax provision" in r["filename"].lower()]
    if not cands:
        return None
    # Prefer the latest version (sort by version string descending).
    cands.sort(key=lambda r: r.get("version") or "", reverse=True)
    return cands[0]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("inputs_json")
    ap.add_argument("pbc_index_json")
    ap.add_argument("out_json")
    args = ap.parse_args()

    inputs = json.loads(Path(args.inputs_json).read_text(encoding="utf-8"))
    pbc_index = json.loads(Path(args.pbc_index_json).read_text(encoding="utf-8"))

    records = []
    fs = extract_fs_tax(inputs)
    # A reported net DTA of "$-" (full valuation allowance) parses to None — treat it
    # as 0 when the deferred table was located so the footing check still runs.
    if fs["net_dta"] is None and fs["total_dta"] is not None:
        fs["net_dta"] = 0.0

    wb_rec = find_provision_wb(pbc_index)
    if not wb_rec:
        print("  (tax provision workbook not found in PBC index — skipping tax module)")
        Path(args.out_json).write_text("[]", encoding="utf-8")
        return

    ref = Path(wb_rec["rel"]).name
    try:
        prov = extract_provision(wb_rec["path"])
    except Exception as e:
        print(f"  WARN: could not read provision workbook {ref}: {e}")
        Path(args.out_json).write_text("[]", encoding="utf-8")
        return

    # Provision values are raw $; FS values are $K. Convert provision to $K.
    p_pretax = prov["pretax"] / 1000.0 if prov["pretax"] is not None else None
    p_stat = prov["statutory_tax"] / 1000.0 if prov["statutory_tax"] is not None else None
    p_total = prov["total_provision"] / 1000.0 if prov["total_provision"] is not None else None
    p_net_dta = prov["net_dta_before_va"] / 1000.0 if prov["net_dta_before_va"] is not None else None
    fed_rate = prov["federal_rate"] or 0.21

    # 1) Statutory recompute: pretax x federal rate (recomputed) vs the workpaper's
    #    "tax at statutory rate". Sign: FS pretax is shown as a loss (negative).
    if p_pretax is not None and p_stat is not None:
        recomputed = p_pretax * fed_rate
        add(records, f"Tax at statutory rate (recompute pretax x {fed_rate:.0%})",
            recomputed, p_stat, ref,
            "Statutory tax recomputes from pretax x federal rate.",
            "Statutory tax does NOT equal pretax x federal rate — check the rate or pretax input.")

    # 2) Rate-rec footing — the workpaper's own statutory->effective check should be ~0.
    rc = prov["recon_check"]
    if rc is not None:
        rc_K = rc / 1000.0
        st = "ties" if abs(rc_K) <= 1.0 else "exception"
        records.append(make_record(
            "tax_provision", pdf_section="FN - Taxes",
            pdf_label="Rate reconciliation foots (statutory -> effective, s/b zero)",
            pdf_year="2025", pdf_value=round(rc_K, 3), source_ref=ref,
            source_label="Reconciliation - s/b zero", source_value=0.0,
            comparison_unit="$K", delta=round(rc_K, 3), tolerance=1.0, status=st,
            is_subtotal=True,
            notes="Rate reconciliation foots to the total provision."
                  if st == "ties" else "Rate reconciliation does NOT foot — investigate.",
        ))

    # 3) Total provision <-> FS income tax expense. FS shows expense as negative; the
    #    workpaper total is a positive expense. Compare magnitudes.
    if p_total is not None and fs["tax_expense"] is not None:
        add(records, "Total income tax expense (provision vs FS)",
            abs(fs["tax_expense"]), abs(p_total), ref,
            "Total provision ties the FS income tax expense.",
            "Total provision does NOT tie the FS income tax expense.")

    # 4) Book pretax reconciliation — provision pretax must equal FS pretax. (The
    #    $X,XXX.XXK class: a footnote/face/provision pretax that drifted apart.)
    if p_pretax is not None and fs["pretax"] is not None:
        add(records, "Book pretax income/(loss) (provision input vs FS)",
            fs["pretax"], p_pretax, ref,
            "Provision is built on the same pretax as the FS.",
            "Provision pretax DIFFERS from the FS pretax — the income statement and the "
            "provision input have drifted apart (the classic IS<->FN tax reconciliation error).")

    # 5) Deferred summary — net DTA before VA ties FS (DTA + DTL); FS table foots to net.
    if p_net_dta is not None and fs["total_dta"] is not None and fs["total_dtl"] is not None:
        fs_net_before_va = fs["total_dta"] + fs["total_dtl"]  # DTL stored negative
        add(records, "Net deferred tax asset before valuation allowance (provision vs FS)",
            fs_net_before_va, p_net_dta, ref,
            "Provision net DTA (before VA) ties the FS deferred tax table.",
            "Provision net DTA (before VA) does NOT tie the FS deferred tax table.")

    # FS deferred table foots: DTA + DTL - VA = net (~0 under a full valuation allowance).
    if all(fs[k] is not None for k in ("total_dta", "total_dtl", "valuation_allowance", "net_dta")):
        computed_net = fs["total_dta"] + fs["total_dtl"] + fs["valuation_allowance"]
        d = computed_net - fs["net_dta"]
        st = "ties" if abs(d) <= 1.0 else "exception"
        records.append(make_record(
            "tax_provision", pdf_section="FN - Taxes",
            pdf_label="FS deferred table foots (DTA + DTL - VA = net DTA)",
            pdf_year="2025", pdf_value=round(computed_net, 1), source_ref="FS FN-07 deferred table",
            source_label="reported net DTA", source_value=round(fs["net_dta"], 1),
            comparison_unit="$K", delta=round(d, 1), tolerance=1.0, status=st,
            is_subtotal=True,
            notes="Deferred tax table foots: total DTA + total DTL - valuation allowance = net."
                  if st == "ties" else "Deferred tax table does NOT foot.",
        ))

    Path(args.out_json).write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")

    by_status = {}
    for r in records:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    print(f"Tax provision module: {len(records)} checks | {by_status}")
    print(f"  Provision: {ref}")
    for r in records:
        flag = "" if r["status"] in ("ties", "ties-with-rounding") else "  <-- REVIEW"
        print(f"  [{r['status']:<10}] {r['pdf_label']}: FS={r['pdf_value']} prov={r['source_value']} d={r['delta']}{flag}")
    print(f"Wrote {args.out_json}")


if __name__ == "__main__":
    main()
