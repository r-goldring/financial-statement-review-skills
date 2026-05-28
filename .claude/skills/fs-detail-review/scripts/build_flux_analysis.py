"""Phase 3 — Analytical / flux review (re-baselined to the FINAL FS).

The company prepares a flux-review PBC with its own thresholds + written comments,
but it is drafted early (the FY25 copy is the Nov-2025 monthly account-level flux),
before the revision rounds — so it goes STALE and will NOT tie the final FS. This
module therefore treats it as a prior draft to **re-baseline**, not a deliverable to
tie out:

  1. Recompute YoY movements on the CURRENT (final) FS — IS + BS, 2025 vs 2024,
     from inputs.json — and apply a $ / % threshold to find the material movers.
  2. Harvest the flux PBC's account-level comments into a directional **comment bank**
     (clearly marked as predating the final FS — context, not a tie).
  3. For each material FS mover: carry a related prior comment if one is found, add a
     suggested driver from references/company-context.template.md, and flag movers with no
     explanation. This automates refreshing the early flux to the final numbers.

Output: a "Flux Review" tab (lane "flux"): line, 2025, 2024, $ var, % var, threshold
breach, carried comment (+ staleness), suggested driver, status.

CLI:
  python build_flux_analysis.py <inputs.json> <pbc-index.json> <out.json>
       [--dollar-threshold 1000] [--pct-threshold 0.10] [--context <company-context.template.md>]
"""

import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from tie_out_common import parse_value, make_record  # noqa: E402

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


# FS lines that are subtotals/headers — recompute but don't demand a comment.
SUBTOTAL_LINES = {
    "gross profit", "total operating expenses", "loss from operations",
    "income from operations", "total other expense", "total other income (expense)",
    "loss before income taxes", "net loss", "comprehensive loss",
    "total current assets", "total assets", "total current liabilities",
    "total liabilities", "total members' equity", "total members equity",
    "total liabilities and members' equity", "total deferred tax assets",
}

# Keywords -> driver hints, harvested from company-context.template.md "Specific FY25 movements".
# These are *suggested* explanations for the recomputed movers, not assertions.
def load_context_drivers(context_path):
    drivers = []
    if not context_path or not Path(context_path).exists():
        return drivers
    text = Path(context_path).read_text(encoding="utf-8")
    # Pull the bullet lines under "Specific FY25 movements observed".
    m = re.search(r"Specific FY25 movements observed(.*?)(?:\n##|\Z)", text, re.S)
    blob = m.group(1) if m else text
    for line in blob.splitlines():
        line = line.strip(" -*\t")
        if len(line) > 25 and any(ch.isdigit() for ch in line):
            drivers.append(line)
    return drivers


def _table_rows(t):
    return t if isinstance(t, list) else t.get("rows", [])


def extract_fs_lines(inputs):
    """Return [(statement, label, cy, py)] for IS + BS lines (values in $K)."""
    tables = inputs.get("fy25_docx", {}).get("tables", [])
    out = []
    for t in tables:
        rows = _table_rows(t)
        flat = " ".join(str(c) for r in rows for c in (r if isinstance(r, list) else []) if c).lower()
        is_bs = "total assets" in flat and "cash" in flat
        is_is = "loss before income taxes" in flat and "revenue" in flat
        if not (is_bs or is_is):
            continue
        statement = "Balance Sheet" if is_bs else "Income Statement"
        for r in rows:
            cells = r if isinstance(r, list) else []
            label = next((str(c) for c in cells if isinstance(c, str) and c.strip()), "")
            if not label or label.lower() in ("2025", "2024", "(in thousands)"):
                continue
            nums = [parse_value(c) for c in cells]
            nums = [n for n in nums if n is not None]
            if len(nums) >= 2:
                out.append((statement, label.strip(), nums[0], nums[1]))
    return out


def harvest_comment_bank(path):
    """From the flux PBC, collect (account_label, comment) pairs across all tabs.
    The comment is the trailing text cell on a row that also has a numeric account."""
    import openpyxl
    bank = []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for tab in wb.sheetnames:
        ws = wb[tab]
        for row in ws.iter_rows(values_only=True):
            cells = list(row)
            label = next((str(c) for c in cells[:2] if isinstance(c, str) and c.strip()), "")
            if not label:
                continue
            texts = [str(c).strip() for c in cells[2:] if isinstance(c, str) and len(str(c).strip()) > 8]
            has_number = any(parse_value(c) is not None for c in cells)
            if texts and has_number:
                bank.append((tab, label.strip(), " | ".join(texts)))
    wb.close()
    return bank


_WORD = re.compile(r"[a-z]{4,}")
_STOP = {"total", "expense", "income", "other", "current", "amount", "subscription",
         "services", "service", "cost", "costs"}


def keywords(s):
    return {w for w in _WORD.findall(s.lower()) if w not in _STOP}


def related_comment(line_label, bank):
    """Find flux-PBC comments whose account label shares a significant keyword with
    the FS line. Returns a joined string (best 2) or None."""
    lk = keywords(line_label)
    if not lk:
        return None
    hits = []
    for tab, acct, comment in bank:
        overlap = lk & keywords(acct + " " + tab)
        if overlap:
            hits.append((len(overlap), f"[{tab}] {acct}: {comment}"))
    if not hits:
        return None
    hits.sort(reverse=True)
    return "  ||  ".join(h[1] for h in hits[:2])


def suggested_driver(line_label, drivers):
    lk = keywords(line_label)
    for d in drivers:
        if lk & keywords(d):
            return d
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("inputs_json")
    ap.add_argument("pbc_index_json")
    ap.add_argument("out_json")
    ap.add_argument("--dollar-threshold", type=float, default=1000.0,
                    help="material $ move in $K (default 1000 = $1.0M)")
    ap.add_argument("--pct-threshold", type=float, default=0.10,
                    help="material %% move (default 0.10 = 10%%)")
    ap.add_argument("--context", default=str(Path(__file__).parent.parent / "references" / "company-context.template.md"))
    args = ap.parse_args()

    inputs = json.loads(Path(args.inputs_json).read_text(encoding="utf-8"))
    pbc_index = json.loads(Path(args.pbc_index_json).read_text(encoding="utf-8"))

    fs_lines = extract_fs_lines(inputs)
    drivers = load_context_drivers(args.context)

    # Locate + harvest the flux PBC (the company's own flux/recon deliverable).
    # Prefer the Year-End annual flux (the one sent to the auditors) over the interim copy.
    flux_cands = [r for r in pbc_index if r.get("category") == "flux"
                  and r.get("ext") in (".xlsx", ".xlsm")
                  and "flux" in r["filename"].lower()]
    flux = next((r for r in flux_cands if "year end" in r["rel"].lower()), None) \
        or (flux_cands[0] if flux_cands else None)
    bank, flux_ref, flux_note = [], None, ""
    if flux:
        flux_ref = Path(flux["rel"]).name
        try:
            bank = harvest_comment_bank(flux["path"])
            flux_note = (f"Prior comments carried from {flux_ref} (monthly account-level flux, "
                         f"drafted before the final FS — directional context, NOT a tie).")
        except Exception as e:
            print(f"  WARN: could not read flux PBC {flux_ref}: {e}")

    records = []
    for statement, label, cy, py in fs_lines:
        var = cy - py
        pct = (var / abs(py)) if py else (1.0 if var else 0.0)
        is_sub = label.lower() in SUBTOTAL_LINES
        breach = abs(var) >= args.dollar_threshold or (abs(pct) >= args.pct_threshold and abs(var) >= 250)

        comment = related_comment(label, bank) if (breach and not is_sub) else None
        driver = suggested_driver(label, drivers) if (breach and not is_sub) else None

        if not breach:
            status = "immaterial"
            note = "Movement below threshold."
        elif is_sub:
            status = "mover-subtotal"
            note = "Material subtotal movement (explained by its components)."
        elif comment or driver:
            status = "mover-explained"
            parts = []
            if comment:
                parts.append(flux_note + "  " + comment)
            if driver:
                parts.append("Suggested driver (from FY25 context): " + driver)
            note = "  ".join(parts)
        else:
            status = "mover-needs-explanation"
            note = ("Material YoY move with no carried comment or context driver — draft an "
                    "explanation. (The flux PBC predates the final FS, so a refreshed comment is "
                    "expected here.)")

        records.append(make_record(
            "flux",
            pdf_section=statement, pdf_label=label, pdf_year="2025 vs 2024",
            pdf_value=cy, source_ref=flux_ref or "(no flux PBC)",
            source_label="prior-year", source_value=py,
            comparison_unit="$K", delta=round(var, 1),
            tolerance=args.dollar_threshold,
            status=status, is_subtotal=is_sub, notes=note,
        ))

    Path(args.out_json).write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")

    by = {}
    for r in records:
        by[r["status"]] = by.get(r["status"], 0) + 1
    print(f"Flux review: {len(records)} FS lines | {by}")
    print(f"  Flux PBC: {flux_ref or '(none found)'} | comment bank: {len(bank)} entries | "
          f"context drivers: {len(drivers)}")
    print(f"  Thresholds: |$|>= {args.dollar_threshold}K  or  |%|>= {args.pct_threshold:.0%}")
    for r in records:
        if r["status"] in ("mover-needs-explanation", "mover-explained"):
            tag = "NEEDS COMMENT" if r["status"] == "mover-needs-explanation" else "explained"
            print(f"  [{tag:<13}] {r['pdf_section'][:2]} {r['pdf_label'][:34]:<34} "
                  f"{r['pdf_value']:>10} vs {r['source_value']:>10}  d={r['delta']:>9}")
    print(f"Wrote {args.out_json}")


if __name__ == "__main__":
    main()
