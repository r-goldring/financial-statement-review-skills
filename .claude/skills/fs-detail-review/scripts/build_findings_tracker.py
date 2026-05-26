"""Build the Findings Tracker Excel workbook from a list of finding records.

Usage:
    python build_findings_tracker.py --findings findings.json --out tracker.xlsx
    python build_findings_tracker.py --merge prior.xlsx --findings new.json --out tracker.xlsx

Carry-over rules (when --merge is supplied):
  - Open / In Discussion findings from prior tracker → re-evaluated against new findings
  - If prior finding's underlying issue still appears in new findings (matched by category + tab + cell + key text) → carry forward, update bridge/PDF version
  - If no longer reproducible → mark Status=Resolved, Resolved In Version, Resolved Date, Resolution Notes
  - Resolved / Deferred / Not Applicable findings → kept as-is

Schema: see references/findings-schema.md
"""

import argparse
import json
from datetime import date
from pathlib import Path

COLUMNS = [
    "Finding ID",
    "Date Identified",
    "Bridge Version",
    "PDF Version",
    "FS Area",
    "Tab",
    "Cell/Range",
    "Severity",
    "Category",
    "Description",
    "Evidence",
    "Recommended Action",
    "Status",
    "Resolution Notes",
    "Resolved In Version",
    "Resolved Date",
]

SEVERITY_FILL = {
    "Critical": "FFCCCC",
    "Material": "FFE5B4",
    "Minor": "FFFACD",
    "Informational": "E0F2FE",
}

STATUS_RESOLVED = {"Resolved", "Deferred", "Not Applicable"}


def _next_id(existing_ids, audit_year=2025):
    used_nums = []
    prefix = f"FY{str(audit_year)[-2:]}-"
    for fid in existing_ids:
        if fid.startswith(prefix):
            try:
                used_nums.append(int(fid[len(prefix):]))
            except ValueError:
                pass
    n = max(used_nums, default=0) + 1
    return f"{prefix}{n:04d}"


def _load_prior_tracker(path):
    """Return list-of-dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = list(rows[0])
    out = []
    for row in rows[1:]:
        if all(c is None or c == "" for c in row):
            continue
        d = {headers[i]: row[i] for i in range(len(headers))}
        out.append(d)
    wb.close()
    return out


def merge_findings(prior, new_findings, bridge_version, pdf_version, audit_year=2025):
    """Merge new findings into a prior tracker dataset. Returns combined list."""
    merged = []
    used_ids = {f.get("Finding ID") for f in prior if f.get("Finding ID")}

    def _key(f):
        return (
            (f.get("Category") or "").strip().lower(),
            (f.get("Tab") or "").strip().lower(),
            (f.get("Cell/Range") or "").strip().lower(),
            (f.get("Description") or "")[:80].strip().lower(),
        )

    new_keyed = {_key(f): f for f in new_findings}

    for prior_f in prior:
        status = (prior_f.get("Status") or "Open").strip()
        if status in STATUS_RESOLVED:
            merged.append(prior_f)
            continue
        # Open or In Discussion — re-evaluate
        if _key(prior_f) in new_keyed:
            updated = dict(prior_f)
            updated["Bridge Version"] = bridge_version
            updated["PDF Version"] = pdf_version
            merged.append(updated)
            new_keyed.pop(_key(prior_f))
        else:
            resolved = dict(prior_f)
            resolved["Status"] = "Resolved"
            resolved["Resolved In Version"] = bridge_version
            resolved["Resolved Date"] = str(date.today())
            resolved["Resolution Notes"] = (
                (prior_f.get("Resolution Notes") or "")
                + " | Auto-detected resolution; verify"
            ).strip(" |")
            merged.append(resolved)

    # Append truly new findings
    for f in new_keyed.values():
        nf = dict(f)
        if not nf.get("Finding ID"):
            new_id = _next_id(used_ids, audit_year)
            nf["Finding ID"] = new_id
            used_ids.add(new_id)
        if not nf.get("Date Identified"):
            nf["Date Identified"] = str(date.today())
        nf["Bridge Version"] = bridge_version
        nf["PDF Version"] = pdf_version
        if not nf.get("Status"):
            nf["Status"] = "Open"
        merged.append(nf)

    return merged


def write_tracker(findings, out_path):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Findings"

    # Header
    header_font = Font(bold=True)
    for c, h in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    # Rows
    for r, f in enumerate(findings, 2):
        for c, h in enumerate(COLUMNS, 1):
            ws.cell(r, c, f.get(h, ""))
        sev = (f.get("Severity") or "").strip()
        if sev in SEVERITY_FILL:
            for c in range(1, len(COLUMNS) + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=SEVERITY_FILL[sev])
        if (f.get("Status") or "").strip() == "Resolved":
            for c in range(1, len(COLUMNS) + 1):
                cell = ws.cell(r, c)
                cell.font = Font(strike=True, color="888888")

    # Column widths
    widths = {
        "Finding ID": 12,
        "Date Identified": 13,
        "Bridge Version": 28,
        "PDF Version": 28,
        "FS Area": 10,
        "Tab": 22,
        "Cell/Range": 14,
        "Severity": 13,
        "Category": 16,
        "Description": 70,
        "Evidence": 60,
        "Recommended Action": 50,
        "Status": 14,
        "Resolution Notes": 40,
        "Resolved In Version": 28,
        "Resolved Date": 13,
    }
    for c, h in enumerate(COLUMNS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = widths.get(h, 18)

    # Wrap text on Description / Evidence
    for r in range(2, ws.max_row + 1):
        for h in ("Description", "Evidence", "Recommended Action", "Resolution Notes"):
            ws.cell(r, COLUMNS.index(h) + 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 60

    # Auto-filter
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(COLUMNS))}{ws.max_row}"

    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--findings", required=True, help="JSON file with list of finding dicts")
    ap.add_argument("--merge", help="Optional prior Findings Tracker.xlsx to merge with")
    ap.add_argument("--bridge-version", required=True)
    ap.add_argument("--pdf-version", default="")
    ap.add_argument("--audit-year", type=int, default=2025)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    new_findings = json.loads(Path(args.findings).read_text())
    prior = _load_prior_tracker(args.merge) if args.merge else []
    merged = merge_findings(prior, new_findings, args.bridge_version, args.pdf_version, args.audit_year)
    write_tracker(merged, args.out)
    print(f"Wrote {args.out} with {len(merged)} findings")
    by_severity = {}
    for f in merged:
        sev = (f.get("Severity") or "").strip()
        by_severity[sev] = by_severity.get(sev, 0) + 1
    print(f"By severity: {by_severity}")


if __name__ == "__main__":
    main()
