"""POC — how much of the tax provision can we recreate from history?

Ingests the last 3 years of the provision workbook (same preparer, same template) and
answers, with evidence, three questions:

  1. CLASSIFY — which rate-rec reconciling items recur every year (mechanical /
     predictable) vs. appear sporadically (judgment / event-driven)? This is the
     "what % is automatable" measure.
  2. TAX-EFFECTING — is each item tax-effected at a stable rate (so the dollar tax
     effect is a deterministic function of the book amount)?
  3. RECOMPUTE — re-derive the current-year permanent-difference tax effects from the
     book amounts x the learned rate, and compare to the preparer's workbook. Deltas
     near zero => the mechanical layer reproduces the preparer.

Plus a multi-year deferred-balance trend (anomaly baseline) from the DTA/DTL summary.

This is a feasibility probe, not a wired lane. It writes nothing into the tie-out.

CLI:
  python tax_provision_recompute_poc.py [--fy23 <xlsx>] [--fy24 <xlsx>] [--fy25 <xlsx>]
"""

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from tie_out_common import parse_value  # noqa: E402

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

ROOT = Path(r"C:\path\to\financial-statement-review")
DEFAULTS = {
    "FY23": ROOT / "Prior Year Examples/2023/FS Compilation Requests (FS Compilation Partner)/8. (YE) Acme Corp 2023 Tax Provision_v2024.5.23.xlsx",
    "FY24": ROOT / "PBCs/2024 Audit/FS Compilation Requests (FS Compilation Partner)/Tax Provision/2024 Acme Corp Tax Provision_v2025.5.27.xlsx",
    "FY25": ROOT / "PBCs/2025 Audit/Year End/FS Compilation Requests (FS Compilation Partner) (old)/9. Tax Provision/9. Acme Corp 2025 Tax Provision_v2026.5.21.xlsx",
}


def _rows(wb, name):
    return [list(r) for r in wb[name].iter_rows(values_only=True)]


def extract_rate_rec(wb):
    """Return {item_label: (book_amount, tax_effect)} for every reconciling line, plus
    pretax, federal_rate, total_provision, recon_check."""
    rows = _rows(wb, "6| Rate Rec.")
    items = {}
    pretax = fed_rate = total = recon = None
    for row in rows:
        # label can be in col 1 (section) or col 2 (perm detail)
        label = None
        for c in (row[1] if len(row) > 1 else None, row[2] if len(row) > 2 else None):
            if isinstance(c, str) and c.strip():
                label = c.strip()
                break
        if not label:
            continue
        ll = label.lower()
        nums = [parse_value(c) for c in row]
        nums = [n for n in nums if n is not None]
        if "pre-tax book income" in ll and nums:
            pretax = nums[0]
        elif "tax at statutory rate" in ll and nums:
            total_stat = nums[0]
            # the rate is the fractional value on the row
            fed_rate = next((n for n in nums if 0 < abs(n) < 1), None)
            items["Tax at Statutory Rate"] = (pretax, total_stat)
        elif "per account rollforward" in ll and nums:
            total = nums[0]
        elif "reconciliation - s/b zero" in ll and nums:
            recon = nums[0]
        else:
            # A reconciling line: book amount (col 3) and/or tax effect.
            book = parse_value(row[3]) if len(row) > 3 else None
            # tax effect is the value at col 4 (perm detail) or the first number for
            # section-level adjustment rows (True-up, DTA Adjustment, etc.).
            te = parse_value(row[4]) if len(row) > 4 else None
            if te is None and nums:
                te = nums[0]
            if (book is not None or te is not None) and ll not in (
                    "permanent differences", "state taxes", "tax provision",
                    "rate reconciliation", "current expense", "deferred provision (bene",
                    "federal benefit"):
                items[label] = (book, te)
    return {"items": items, "pretax": pretax, "federal_rate": fed_rate,
            "total_provision": total, "recon_check": recon}


def extract_dta_balances(wb):
    """{difference_label: total_balance} from '4| DTA DTL Summary' detail rows (those
    with a single-letter footnote group code in col 6)."""
    if "4| DTA DTL Summary" not in wb.sheetnames:
        return {}
    out = {}
    for row in _rows(wb, "4| DTA DTL Summary"):
        label = row[1] if len(row) > 1 else None
        val = parse_value(row[3]) if len(row) > 3 else None
        code = row[6] if len(row) > 6 else None
        if (isinstance(label, str) and label.strip() and val is not None
                and isinstance(code, str) and len(code.strip()) == 1 and code.strip().isalpha()):
            out[label.strip()] = val
    return out


def main():
    ap = argparse.ArgumentParser()
    for yr, p in DEFAULTS.items():
        ap.add_argument(f"--{yr.lower()}", default=str(p))
    args = ap.parse_args()

    import openpyxl
    years = {}
    dta = {}
    for yr in ("FY23", "FY24", "FY25"):
        path = getattr(args, yr.lower())
        if not Path(path).exists():
            print(f"  (missing {yr}: {path})")
            continue
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        years[yr] = extract_rate_rec(wb)
        dta[yr] = extract_dta_balances(wb)
        wb.close()

    yrs = list(years.keys())
    print(f"\n=== Tax provision recompute POC — {', '.join(yrs)} ===")
    for yr in yrs:
        d = years[yr]
        print(f"  {yr}: pretax {d['pretax']:,.0f} | fed rate {d['federal_rate']} | "
              f"total provision {d['total_provision']:,.0f} | recon check {d['recon_check']:,.1f}")

    # ---- 1) CLASSIFY: recurring vs sporadic reconciling items ----
    all_items = []
    for yr in yrs:
        for k in years[yr]["items"]:
            if k not in all_items:
                all_items.append(k)

    def nonzero(yr, item):
        v = years[yr]["items"].get(item)
        if not v:
            return False
        book, te = v
        return (book not in (None, 0)) or (te not in (None, 0))

    print("\n--- 1) Reconciling items: recurrence across years ---")
    print(f"  {'item':<32} " + " ".join(f"{y:>14}" for y in yrs) + "   class")
    recurring = []
    for item in all_items:
        present = [nonzero(yr, item) for yr in yrs]
        cells = []
        for yr in yrs:
            v = years[yr]["items"].get(item)
            te = v[1] if v else None
            cells.append(f"{te:>14,.0f}" if isinstance(te, (int, float)) else f"{'-':>14}")
        cls = "RECURRING" if all(present) else ("occasional" if any(present) else "zero")
        if all(present):
            recurring.append(item)
        print(f"  {item[:32]:<32} " + " ".join(cells) + f"   {cls}")

    # ---- 2) TAX-EFFECTING: implied rate per perm item per year ----
    print("\n--- 2) Tax-effecting rate (tax_effect / book_amount) — perms only ---")
    print(f"  {'item':<32} " + " ".join(f"{y:>8}" for y in yrs))
    rate_stable = 0
    rate_checked = 0
    for item in recurring:
        rates = []
        for yr in yrs:
            v = years[yr]["items"].get(item)
            if v and v[0] not in (None, 0) and v[1] is not None:
                rates.append(v[1] / v[0])
            else:
                rates.append(None)
        shown = " ".join(f"{r:>8.1%}" if r is not None else f"{'-':>8}" for r in rates)
        valid = [r for r in rates if r is not None]
        if len(valid) >= 2:
            rate_checked += 1
            if max(valid) - min(valid) <= 0.01:  # within 1pt => stable
                rate_stable += 1
        print(f"  {item[:32]:<32} {shown}")

    # ---- 3) RECOMPUTE FY25 perms: book x federal rate vs preparer ----
    fy = "FY25" if "FY25" in years else yrs[-1]
    fed = years[fy]["federal_rate"] or 0.21
    print(f"\n--- 3) Recompute {fy} permanent-difference tax effects (book x {fed:.0%}) vs preparer ---")
    print(f"  {'item':<32} {'book':>16} {'recomputed':>14} {'preparer':>14} {'delta':>12}")
    reproduced = 0
    perm_count = 0
    for item in recurring:
        v = years[fy]["items"].get(item)
        if not v or v[0] in (None, 0) or v[1] is None:
            continue
        book, te = v
        recomputed = book * fed
        delta = recomputed - te
        # only count "pure perm" items (tax-effected ~ at the federal rate)
        if abs(te / book - fed) <= 0.02:
            perm_count += 1
            if abs(delta) <= 1.0:
                reproduced += 1
        print(f"  {item[:32]:<32} {book:>16,.0f} {recomputed:>14,.0f} {te:>14,.0f} {delta:>12,.1f}")

    # ---- 4) Multi-year deferred-balance trend (anomaly baseline) ----
    print("\n--- 4) Deferred balances by difference (DTA/DTL summary), 3-year trend ---")
    keys = []
    for yr in yrs:
        for k in dta.get(yr, {}):
            if k not in keys:
                keys.append(k)
    print(f"  {'difference':<30} " + " ".join(f"{y:>16}" for y in yrs))
    for k in keys:
        cells = " ".join(f"{dta[yr].get(k):>16,.0f}" if dta.get(yr, {}).get(k) is not None
                         else f"{'-':>16}" for yr in yrs)
        print(f"  {k[:30]:<30} {cells}")

    # ---- Verdict ----
    print("\n=== FEASIBILITY VERDICT ===")
    print(f"  Reconciling items seen: {len(all_items)} | recurring all {len(yrs)} yrs: {len(recurring)} "
          f"({len(recurring)/len(all_items):.0%})")
    if rate_checked:
        print(f"  Tax-effecting rate stable (<=1pt drift): {rate_stable}/{rate_checked} recurring items")
    if perm_count:
        print(f"  {fy} permanent diffs reproduced within $1K by (book x {fed:.0%}): {reproduced}/{perm_count}")
    print("  => The recurring + rate-stable items are the mechanical layer (auto-recompute, then check).")
    print("     The 'occasional' items + non-federal-rate adjustments are the judgment/event layer.")


if __name__ == "__main__":
    main()
