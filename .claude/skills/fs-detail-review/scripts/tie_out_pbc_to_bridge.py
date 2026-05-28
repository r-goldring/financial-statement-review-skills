"""Lane 8 — PBC -> bridge tie-out (source-to-disclosure assurance).

Closes the last integrity gap: today the bridge's footnote numbers are taken on
faith. This lane ties each mapped footnote/bridge tab back to the *supporting PBC
workpaper* (and, where the PBC carries a GL column, to the general ledger), so a
stale compiler schedule that no longer agrees with the company's records is caught.

The flagship check is the **3-way intangibles reconcile** (the FY25 finding we hit
by hand): the compiler's bridge FN-Intangibles tab vs. the company's
Goodwill & Intangibles rollforward (subledger) vs. the GL column on that rollforward.
When **net ties but gross and accumulated amortization each diverge by the same
amount (offsetting)**, that's the fully-amortized-asset gotcha — a real exception,
not noise. See references/source-of-truth-hierarchy.md.

Records are emitted under lane "pbc_to_bridge" and rendered in the "PBC Ties" tab.

CLI:
  python tie_out_pbc_to_bridge.py <inputs.json> <pbc-index.json> <out.json>
"""

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from tie_out_common import parse_value, compare, make_record  # noqa: E402

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────────────
# Bridge-tab reading (from inputs.json bridge.tabs[*].rows)
# ─────────────────────────────────────────────────────────────────────────────
def get_bridge_tab(inputs, tab_name):
    for t in inputs.get("bridge", {}).get("tabs", []):
        if t["name"] == tab_name:
            return t
    return None


def find_section_row(rows, header_substrings, start=0):
    """Return the row index whose joined text contains all header_substrings."""
    for i in range(start, len(rows)):
        joined = " ".join(str(c) for c in rows[i] if c is not None).lower()
        if all(h.lower() in joined for h in header_substrings):
            return i
    return None


def find_labeled_row(rows, label_substr, start=0, end=None, require_min_numerics=0):
    """Return (row_index, list_of_numeric_values) for the first row whose 2nd cell
    (label col) contains label_substr and has at least require_min_numerics numbers."""
    end = end if end is not None else len(rows)
    for i in range(start, min(end, len(rows))):
        row = rows[i]
        label = next((str(c) for c in row[:3] if isinstance(c, str) and c.strip()), "")
        if label_substr.lower() in label.lower():
            nums = [parse_value(c) for c in row]
            nums = [n for n in nums if n is not None]
            if len(nums) >= require_min_numerics:
                return i, row, nums
    return None, None, None


# ─────────────────────────────────────────────────────────────────────────────
# PBC rollforward extraction (Goodwill & Intangibles Summary tab)
# ─────────────────────────────────────────────────────────────────────────────
def extract_gi_rollforward(path):
    """From a 'Goodwill & Intangibles Rollforward' Summary tab, pull ending
    gross / accum / net + the GL column for Intangibles and Goodwill.

    Layout (both the company and compiler versions): per-entity columns (5), then
    Total, then Per-GL, then Variance. Ending rows carry the GL+variance columns;
    opening/activity rows do not — so 'has a GL column' (>=7 numerics) disambiguates
    the ending row from the opening row of the same label.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["1. Summary"] if "1. Summary" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    def ending(label_substr, exclude=()):
        """Find the ending row (the one with a GL column) for a label; return
        (total, gl, variance) in raw dollars. `exclude` rejects rows whose label
        contains any of these terms — needed because the gross row label
        ('Intangibles 12/31/2025') is a substring of the accum row label
        ('Accumulated Amortization - Intangibles 12/31/2025')."""
        best = None
        for row in rows:
            label = next((str(c) for c in row[:3] if isinstance(c, str) and c.strip()), "")
            ll = label.lower()
            if label_substr.lower() not in ll:
                continue
            if any(x.lower() in ll for x in exclude):
                continue
            nums = [parse_value(c) for c in row if parse_value(c) is not None]
            if len(nums) >= 7:  # 5 entities + Total + GL (+ optional variance)
                total = nums[5]
                gl = nums[6]
                variance = nums[7] if len(nums) >= 8 else (total - gl)
                best = (total, gl, variance)
        return best

    return {
        "intangibles_gross": ending("Intangibles 12/31/2025", exclude=("accumulated", "net")),
        "intangibles_accum": ending("Accumulated Amortization - Intangi"),
        "goodwill_gross": ending("Goodwill 12/31/2025", exclude=("accumulated", "net")),
        "goodwill_accum": ending("Accumulated Amortization - Goodwil"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Bridge FN-Intangibles ending totals
# ─────────────────────────────────────────────────────────────────────────────
def extract_bridge_intangibles(inputs):
    """Return bridge FN-Intangibles 'As of December 31, 2025' totals (in $K):
    {intangibles: (gross, accum, net), goodwill: (gross, accum, net)}."""
    tab = get_bridge_tab(inputs, "FN - Intangibles")
    if not tab:
        return None
    rows = tab["rows"]
    # The 2025 section header — pick the LATER 'december 31, 2025' so we don't grab
    # the 2024 block. Columns: gross=5, accum=7, net=9 (0-based) in the work block.
    sec = find_section_row(rows, ["december", "31, 2025"], start=8)
    if sec is None:
        sec = 0

    def total(label_substr):
        idx, row, _ = find_labeled_row(rows, label_substr, start=sec)
        if row is None:
            return None
        return (parse_value(row[5]), parse_value(row[7]), parse_value(row[9]))

    return {
        "intangibles": total("Total intangible assets"),
        "goodwill": total("Total goodwill"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Tie logic
# ─────────────────────────────────────────────────────────────────────────────
def tie_three_way(records, fs_area, line_name, bridge_kv, pbc_triplet, pbc_ref):
    """bridge_kv: (gross_K, accum_K, net_K) from the bridge (compiler schedule, $K).
    pbc_triplet: {'gross': (total,gl,var), 'accum': (...)} in raw $ from the subledger PBC.
    Emits: (a) subledger<->GL internal checks, (b) bridge<->subledger ties, and a
    diagnostic when gross/accum diverge offsetting while net ties."""
    if not bridge_kv or not pbc_triplet:
        return

    b_gross, b_accum, b_net = bridge_kv
    components = [("gross carrying amount", 0, pbc_triplet.get("gross")),
                 ("accumulated amortization", 1, pbc_triplet.get("accum"))]

    gross_delta = accum_delta = None

    for comp_name, bidx, pbc_vals in components:
        if pbc_vals is None:
            continue
        pbc_total, pbc_gl, pbc_var = pbc_vals
        bridge_val = bridge_kv[bidx]

        # (a) subledger <-> GL. Confirms the company's rec ties its ledger. Compared
        # in $K (like the rest of the tooling) so sub-$K rounding on large balances
        # (e.g. a $391 variance on $228M goodwill) is noise, not a false exception.
        d_gl, st_gl, tol_gl, unit_gl = compare(pbc_total / 1000.0, pbc_gl / 1000.0,
                                               "$K", "$K", is_subtotal=True)
        records.append(make_record(
            "pbc_to_bridge",
            pdf_section=fs_area, pdf_label=f"{line_name} — {comp_name} (subledger vs GL)",
            pdf_year="2025", pdf_value=round(pbc_total / 1000.0, 1),
            source_ref=pbc_ref, source_label="Per GL column", source_value=round(pbc_gl / 1000.0, 1),
            comparison_unit=unit_gl, delta=round(d_gl, 1) if d_gl is not None else None,
            tolerance=tol_gl, status=st_gl, is_subtotal=True,
            notes="Subledger rollforward ties to the general ledger (variance column)."
                  if st_gl in ("ties", "ties-with-rounding") else
                  "Subledger does NOT tie to its own GL — investigate the rec before the bridge.",
        ))

        # (b) bridge (compiler schedule, $K) <-> subledger PBC (raw $). The real tie.
        d, st, tol, unit = compare(bridge_val, pbc_total, "$K", "$1", is_subtotal=True)
        if comp_name.startswith("gross"):
            gross_delta = d
        else:
            accum_delta = d
        records.append(make_record(
            "pbc_to_bridge",
            pdf_section=fs_area, pdf_label=f"{line_name} — {comp_name} (bridge vs PBC)",
            pdf_year="2025", pdf_value=bridge_val,
            source_ref=pbc_ref, source_label=f"{line_name} {comp_name} (subledger total)",
            source_value=round(pbc_total / 1000.0, 1),
            comparison_unit=unit, delta=round(d, 1) if d is not None else None,
            tolerance=tol, status=st, is_subtotal=True,
            notes="Bridge footnote ties to the supporting PBC." if st in ("ties", "ties-with-rounding")
                  else "Bridge footnote does NOT tie to the supporting PBC / GL.",
        ))

    # Net check (bridge net vs subledger net = gross+accum).
    pbc_net = None
    if pbc_triplet.get("gross") and pbc_triplet.get("accum"):
        pbc_net = pbc_triplet["gross"][0] + pbc_triplet["accum"][0]  # accum is negative
    if pbc_net is not None and b_net is not None:
        d, st, tol, unit = compare(b_net, pbc_net, "$K", "$1", is_subtotal=True)
        records.append(make_record(
            "pbc_to_bridge",
            pdf_section=fs_area, pdf_label=f"{line_name} — net carrying amount (bridge vs PBC)",
            pdf_year="2025", pdf_value=b_net, source_ref=pbc_ref,
            source_label=f"{line_name} net (subledger)", source_value=round(pbc_net / 1000.0, 1),
            comparison_unit=unit, delta=round(d, 1) if d is not None else None,
            tolerance=tol, status=st, is_subtotal=True,
            notes="Net carrying amount ties." if st in ("ties", "ties-with-rounding")
                  else "Net carrying amount does not tie.",
        ))

    # Diagnostic: net ties but gross & accum diverge offsetting → fully-amortized gotcha.
    if (gross_delta is not None and accum_delta is not None
            and abs(gross_delta) > 5.0 and abs(accum_delta) > 5.0
            and abs(gross_delta + accum_delta) <= 5.0):
        records.append(make_record(
            "pbc_to_bridge",
            pdf_section=fs_area, pdf_label=f"{line_name} — DIAGNOSTIC: offsetting gross/accum overstatement",
            pdf_year="2025", pdf_value=round(gross_delta, 1),
            source_ref=pbc_ref, source_label="accum delta (offsetting)",
            source_value=round(accum_delta, 1),
            comparison_unit="$K", delta=round(gross_delta + accum_delta, 1),
            tolerance=5.0, status="exception", is_subtotal=True,
            notes=(f"Net ties but gross is overstated by ~{abs(gross_delta):,.0f}K and accumulated "
                   f"amortization by ~{abs(accum_delta):,.0f}K (offsetting). Classic fully-amortized / "
                   f"disposed-asset gotcha: the compiler's bridge schedule still carries assets the GL "
                   f"wrote off. Fix the bridge FN tab to match the PBC/GL ending balances. "
                   f"See references/source-of-truth-hierarchy.md."),
        ))


def find_pbc(pbc_index, category, prefer_substr=None):
    """Pick the PBC file for a category. If prefer_substr given, prefer a path
    containing it (e.g. the company's own 'Goodwill (YE)' rollforward as subledger)."""
    cands = [r for r in pbc_index if r.get("category") == category and r.get("fs_relevant")
             and r.get("ext") in (".xlsx", ".xls", ".xlsm")]
    if not cands:
        return None
    if prefer_substr:
        for r in cands:
            if prefer_substr.lower() in r["rel"].lower():
                return r
    return cands[0]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("inputs_json")
    ap.add_argument("pbc_index_json")
    ap.add_argument("out_json")
    args = ap.parse_args()

    inputs = json.loads(Path(args.inputs_json).read_text(encoding="utf-8"))
    pbc_index = json.loads(Path(args.pbc_index_json).read_text(encoding="utf-8"))

    records = []

    # ── Intangibles & goodwill 3-way reconcile (the flagship Lane 8 check) ──────
    # Subledger source: prefer the company's own 'Goodwill (YE)' rollforward; fall
    # back to the compiler's copy in the FS Compilation folder.
    gi = (find_pbc(pbc_index, "goodwill-intangibles", prefer_substr="Goodwill (YE)")
          or find_pbc(pbc_index, "goodwill-intangibles", prefer_substr="Rollforward"))
    bridge_intang = extract_bridge_intangibles(inputs)

    if gi and bridge_intang:
        pbc_ref = Path(gi["rel"]).name
        try:
            rf = extract_gi_rollforward(gi["path"])
        except Exception as e:
            rf = None
            print(f"  WARN: could not read intangibles rollforward {pbc_ref}: {e}")
        if rf:
            tie_three_way(records, "FN - Intangibles", "Total intangible assets",
                          bridge_intang.get("intangibles"),
                          {"gross": rf["intangibles_gross"], "accum": rf["intangibles_accum"]},
                          pbc_ref)
            tie_three_way(records, "FN - Intangibles", "Total goodwill",
                          bridge_intang.get("goodwill"),
                          {"gross": rf["goodwill_gross"], "accum": rf["goodwill_accum"]},
                          pbc_ref)
    else:
        print("  (intangibles rollforward PBC or bridge FN-Intangibles tab not found — skipping)")

    Path(args.out_json).write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")

    # Summary
    by_status = {}
    for r in records:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    print(f"Lane 8 (PBC -> bridge): {len(records)} records | {by_status}")
    for r in records:
        if r["status"] not in ("ties", "ties-with-rounding"):
            print(f"  [{r['status']:<10}] {r['pdf_label']}  bridge={r['pdf_value']} "
                  f"pbc={r['source_value']} delta={r['delta']}")
    print(f"Wrote {args.out_json}")


if __name__ == "__main__":
    main()
