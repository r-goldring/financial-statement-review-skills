"""Robust bridge-workbook extractor.

Strategies, in order of preference:
  1. openpyxl (data_only=True, normal mode) — best fidelity for formulas + named ranges
  2. openpyxl (data_only=True, read_only=True) — for very large workbooks
  3. Direct ZIP+XML parse via openpyxl internals — for files with features openpyxl chokes on
     (the FY25 partial-update bridge exhibited this)

Outputs a normalized JSON structure:
  {
    "path": str,
    "strategy": "openpyxl-normal" | "openpyxl-readonly" | "zip-xml-fallback",
    "sheet_names": [str],
    "sheets": {
      sheet_name: {
        "max_row": int,
        "max_col": int,
        "rows": [[cell_value, ...], ...]   # 2D list, sparse for trailing empties
      }
    },
    "warnings": [str]
  }

CLI:
  python extract_bridge.py <bridge.xlsx>                # prints sheet names
  python extract_bridge.py <bridge.xlsx> --json <out>   # writes full JSON dump
  python extract_bridge.py <bridge.xlsx> --sheet NAME   # prints one sheet as TSV
"""

import argparse
import json
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path


def _try_openpyxl(path, read_only):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=read_only)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        if read_only:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            max_row = len(rows)
            max_col = max((len(r) for r in rows), default=0)
        else:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            rows = []
            for r in range(1, max_row + 1):
                rows.append([ws.cell(r, c).value for c in range(1, max_col + 1)])
        sheets[name] = {"max_row": max_row, "max_col": max_col, "rows": rows}
    wb.close()
    return wb.sheetnames, sheets


def _zip_xml_fallback(path):
    """Parse the .xlsx ZIP directly. Useful when openpyxl chokes on advanced features.

    Reads workbook.xml for sheet names + sheet1.xml etc. for cell values.
    Inline strings, shared strings, and number values are recovered.
    Formulas are not evaluated — cached values from <v> elements are used.
    """
    NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

    with zipfile.ZipFile(path) as zf:
        # 1. Read shared strings table
        shared_strings = []
        if "xl/sharedStrings.xml" in zf.namelist():
            tree = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in tree.findall(f"{NS}si"):
                # Concatenate all <t> descendants (handles rich text)
                parts = [t.text or "" for t in si.iter(f"{NS}t")]
                shared_strings.append("".join(parts))

        # 2. Read workbook.xml for sheet names + sheet IDs
        wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        # Map sheet name → relationship id
        sheets_meta = []
        for s in wb_xml.find(f"{NS}sheets").findall(f"{NS}sheet"):
            sheets_meta.append({
                "name": s.attrib["name"],
                "sheetId": s.attrib["sheetId"],
                "rId": s.attrib[f"{REL}id"],
            })

        # 3. Read workbook rels to map rId → sheet xml path
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
        rel_map = {}
        for r in rels_xml.findall(f"{REL_NS}Relationship"):
            rel_map[r.attrib["Id"]] = r.attrib["Target"]

        # 4. Parse each sheet
        sheet_names = [s["name"] for s in sheets_meta]
        sheets = {}
        for s in sheets_meta:
            target = rel_map[s["rId"]]
            xml_path = f"xl/{target}" if not target.startswith("xl/") else target
            xml_path = xml_path.replace("xl/xl/", "xl/")
            try:
                sheet_xml = ET.fromstring(zf.read(xml_path))
            except KeyError:
                sheets[s["name"]] = {"max_row": 0, "max_col": 0, "rows": [],
                                     "_warning": f"sheet xml not found at {xml_path}"}
                continue

            sheet_data = sheet_xml.find(f"{NS}sheetData")
            if sheet_data is None:
                sheets[s["name"]] = {"max_row": 0, "max_col": 0, "rows": []}
                continue

            cells_by_rc = {}
            max_row = 0
            max_col = 0
            for row in sheet_data.findall(f"{NS}row"):
                r = int(row.attrib.get("r", "0"))
                if r > max_row:
                    max_row = r
                for c in row.findall(f"{NS}c"):
                    ref = c.attrib.get("r", "")  # e.g. "B12"
                    cell_type = c.attrib.get("t", "n")  # default: number
                    v = c.find(f"{NS}v")
                    is_node = c.find(f"{NS}is")
                    value = None
                    if cell_type == "s" and v is not None:
                        try:
                            value = shared_strings[int(v.text)]
                        except (ValueError, IndexError):
                            value = v.text
                    elif cell_type == "inlineStr" and is_node is not None:
                        parts = [t.text or "" for t in is_node.iter(f"{NS}t")]
                        value = "".join(parts)
                    elif cell_type == "b" and v is not None:
                        value = v.text == "1"
                    elif cell_type == "str" and v is not None:
                        value = v.text  # formula result
                    elif v is not None and v.text is not None:
                        try:
                            num = float(v.text)
                            value = int(num) if num.is_integer() else num
                        except ValueError:
                            value = v.text
                    col_letter = "".join(ch for ch in ref if ch.isalpha())
                    col_num = 0
                    for ch in col_letter:
                        col_num = col_num * 26 + (ord(ch.upper()) - ord("A") + 1)
                    if col_num > max_col:
                        max_col = col_num
                    if value is not None:
                        cells_by_rc[(r, col_num)] = value

            rows = []
            for r in range(1, max_row + 1):
                row = [cells_by_rc.get((r, c)) for c in range(1, max_col + 1)]
                rows.append(row)
            sheets[s["name"]] = {"max_row": max_row, "max_col": max_col, "rows": rows}

        return sheet_names, sheets


def extract_bridge(path):
    path = str(path)
    warnings = []

    for strategy_name, fn in [
        ("openpyxl-normal", lambda: _try_openpyxl(path, read_only=False)),
        ("openpyxl-readonly", lambda: _try_openpyxl(path, read_only=True)),
        ("zip-xml-fallback", lambda: _zip_xml_fallback(path)),
    ]:
        try:
            sheet_names, sheets = fn()
            return {
                "path": path,
                "strategy": strategy_name,
                "sheet_names": sheet_names,
                "sheets": sheets,
                "warnings": warnings,
            }
        except Exception as e:
            warnings.append(f"{strategy_name} failed: {type(e).__name__}: {e}")

    raise RuntimeError(
        f"All extraction strategies failed for {path}. Warnings: {warnings}"
    )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("path", help="Path to .xlsx file")
    ap.add_argument("--json", help="Write full JSON dump to this path")
    ap.add_argument("--sheet", help="Print just one sheet as TSV")
    ap.add_argument("--names-only", action="store_true", help="Print sheet names only")
    args = ap.parse_args()

    result = extract_bridge(args.path)
    if args.names_only:
        print("\n".join(result["sheet_names"]))
        return

    if args.sheet:
        if args.sheet not in result["sheets"]:
            print(f"Sheet '{args.sheet}' not found. Available:", file=sys.stderr)
            print("\n".join(result["sheet_names"]), file=sys.stderr)
            sys.exit(1)
        sheet = result["sheets"][args.sheet]
        for row in sheet["rows"]:
            print("\t".join("" if v is None else str(v) for v in row))
        return

    if args.json:
        # Full dump can be large — round numerics to keep size sane
        Path(args.json).write_text(json.dumps(result, default=str, indent=2))
        print(f"Wrote {args.json}")
        return

    # Default: print summary
    print(f"File: {result['path']}")
    print(f"Strategy used: {result['strategy']}")
    print(f"Sheets ({len(result['sheet_names'])}):")
    for name in result["sheet_names"]:
        s = result["sheets"][name]
        print(f"  {name}  ({s['max_row']}x{s['max_col']})")
    if result["warnings"]:
        print("\nWarnings:")
        for w in result["warnings"]:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
