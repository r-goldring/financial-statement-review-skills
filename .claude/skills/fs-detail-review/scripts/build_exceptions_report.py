"""Build an Excel tie-out exceptions report.

Reads tie-records JSON files and produces a single Excel workbook with one tab:
  - 'Exceptions' — every record where status != 'ties' / 'ties-with-rounding' /
                  'ties-with-sign-inversion'
  - 'All Records' (optional, --include-all) — every record, for audit trail

Columns:
  Lane | PDF Page | Section | Label | Year | PDF Value | Source Ref |
  Source Label | Source Value | Comparison Unit | Delta | Tolerance | Status |
  Is Subtotal | Notes

CLI:
  python build_exceptions_report.py <out.xlsx> <tie_records.json> [<tie_records.json> ...]
       [--include-all]
"""

import argparse
import json
import re
import sys
from pathlib import Path

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


# The canonical FS-build categories. The register flags any of these that yields
# zero files as MISSING (a PBC that should be on hand before fieldwork but isn't).
EXPECTED_FS_BUILD_CATEGORIES = {
    "tax": "Income taxes (FN-07)",
    "goodwill-intangibles": "Goodwill & intangibles (FN-03/04)",
    "cap-software": "Capitalized software (FN-03)",
    "cap-commissions": "Deferred commissions",
    "lease": "Leases (FN-11)",
    "debt": "Term loans & LOC (FN-06)",
    "401k": "Employee benefit plan (FN-13)",
    "ebitda": "Non-GAAP / EBITDA",
    "going-concern": "Going concern (FN-01)",
    "flux": "Analytical / flux review",
    "trial-balance": "Trial balance (source)",
    "gl-detail": "General ledger (source)",
    "ppe": "Property & equipment (FN-03)",
}

# A "prep version" date (vYYYY.M.D / _recon_YYYY-MM) signals when the workpaper was last
# revised. An as-of date (12.31.25) is the period end and is NOT a staleness signal.
PREP_VERSION_RE = re.compile(r"v?(20\d{2})[._-](\d{1,2})[._-](\d{1,2})", re.I)
RECON_PREP_RE = re.compile(r"_recon_(20\d{2})-(\d{2})", re.I)


def parse_prep_date(version_str):
    """Return (yyyy, mm, dd) tuple for a prep-version date, or None for as-of/none."""
    if not version_str:
        return None
    m = PREP_VERSION_RE.search(version_str)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        # vYYYY.M.D prep dates are in the close/audit window (year > period-end year).
        return (y, mo, d)
    m = RECON_PREP_RE.search(version_str)
    if m:
        return (int(m.group(1)), int(m.group(2)), 1)
    return None


COLUMNS = [
    ("Lane", "lane"),
    ("PDF Page", "pdf_page"),
    ("Section", "pdf_section"),
    ("Label", "pdf_label"),
    ("Year", "pdf_year"),
    ("PDF Value", "pdf_value"),
    ("Source Ref", "source_ref"),
    ("Source Label", "source_label"),
    ("Source Value", "source_value"),
    ("Comparison Unit", "comparison_unit"),
    ("Delta", "delta"),
    ("Tolerance", "tolerance"),
    ("Status", "status"),
    ("Subtotal?", "is_subtotal"),
    ("Notes", "notes"),
]

TIE_STATUSES = {"ties", "ties-with-rounding", "ties-with-sign-inversion",
                "ties-F", "ties-xF", "ties-caption-changed", "ties-no-tb-needed",
                "ties-recomputed", "ties-rate-only"}

# Lane 10 (flux) is analytical, not a tie-out — its statuses are not tie/exception and
# must not pollute the Exceptions tab or the exception counts.
ANALYTICAL_STATUSES = {"immaterial", "mover-explained", "mover-needs-explanation",
                       "mover-subtotal",
                       # deferred-tax cross-year trend rows (informational, not tie/exception)
                       "deferred-trend", "deferred-dropped", "deferred-new",
                       # NOL trend rows (informational)
                       "nol-trend", "nol-decrease"}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("out", help="Output .xlsx")
    ap.add_argument("tie_records", nargs="+", help="One or more tie-record JSONs")
    ap.add_argument("--include-all", action="store_true",
                    help="Include a second 'All Records' tab with every record")
    ap.add_argument("--pbc-index", default=None,
                    help="Path to .work/pbc-index.json — adds a 'PBC Register' tab")
    ap.add_argument("--bridge-asof", default="2026-05-01",
                    help="Final-bridge cutoff (YYYY-MM-DD). FS-build PBCs whose prep "
                         "version predates this are flagged possibly-stale for re-confirmation.")
    args = ap.parse_args()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    all_records = []
    for p in args.tie_records:
        recs = json.loads(Path(p).read_text(encoding="utf-8"))
        all_records.extend(recs)

    print(f"Loaded {len(all_records)} records from {len(args.tie_records)} files")

    # Status summary
    summary = {}
    for r in all_records:
        summary[r["status"]] = summary.get(r["status"], 0) + 1
    print(f"Status summary: {summary}")

    exceptions = [r for r in all_records
                  if r["status"] not in TIE_STATUSES and r["status"] not in ANALYTICAL_STATUSES]
    print(f"Exceptions (non-tie): {len(exceptions)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exceptions"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D52027", end_color="D52027", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Group/sort exceptions: by Lane, then Section, then Page, then Label
    lane_order = {
        "pdf_to_bridge": 1,
        "bridge_to_tb": 2,
        "pdf_prior_year": 3,
        "pdf_internal": 4,
        "soe": 5,
        "footing": 6,
        "mapping_completeness": 7,
        "pbc_to_bridge": 8,
        "tax_provision": 9,
        "flux": 10,
        "tax_recompute": 11,
        "deferred_tax": 12,
        "current_nol": 13,
        "rate_rec": 14,
        "state_tax": 15,
    }
    exceptions.sort(key=lambda r: (
        lane_order.get(r["lane"], 99),
        r["pdf_section"] or "",
        r["pdf_page"] or 0,
        r["pdf_label"] or "",
        r["pdf_year"] or "",
    ))

    for r_idx, rec in enumerate(exceptions, start=2):
        for c_idx, (_, field) in enumerate(COLUMNS, start=1):
            v = rec.get(field)
            if field == "pdf_value" or field == "source_value" or field == "delta" or field == "tolerance":
                if v is not None:
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False if field != "notes" and field != "pdf_label" else True)
            if field == "status":
                if v == "exception":
                    cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                    cell.font = Font(bold=True, color="A52022")
                elif v == "restatement":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif v and "missing" in v:
                    cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Column widths
    widths = [14, 8, 10, 50, 8, 14, 24, 30, 14, 8, 12, 10, 18, 10, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # Optional: All Records tab
    if args.include_all:
        ws_all = wb.create_sheet("All Records")
        for c, (label, _) in enumerate(COLUMNS, start=1):
            cell = ws_all.cell(row=1, column=c, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        for r_idx, rec in enumerate(all_records, start=2):
            for c_idx, (_, field) in enumerate(COLUMNS, start=1):
                v = rec.get(field)
                cell = ws_all.cell(row=r_idx, column=c_idx, value=v)
                cell.border = border
        for i, w in enumerate(widths, start=1):
            ws_all.column_dimensions[get_column_letter(i)].width = w
        ws_all.freeze_panes = "A2"

    # ==== PY Ties tab — every FY25 prior-year (2024) column tie to FY24 final ====
    py_records = [r for r in all_records if r.get("lane") == "pdf_prior_year"]
    ws_py = wb.create_sheet("PY Ties", 1)  # insert as second tab
    ws_py["A1"] = "Prior-Year (FY24 col) Ties — FY25 PDF 2024 column ↔ FY24 final FS"
    ws_py["A1"].font = Font(bold=True, size=13)
    ws_py["A2"] = (
        "'ties' = exact match  |  'ties-caption-changed' = number matches but FY24 caption differs  |  "
        "'ties-with-sign-inversion' = magnitude matches, sign convention differs  |  "
        "'restatement' = number differs (TRUE exception)  |  "
        "'missing-on-fy24-pdf' = no FY24 line matched (likely caption change or new disclosure)"
    )
    ws_py["A2"].font = Font(italic=True, color="606060")
    ws_py.merge_cells("A2:O2")
    ws_py.row_dimensions[2].height = 30
    py_header_row = 4
    for c, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws_py.cell(row=py_header_row, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Sort: real exceptions first (restatement + missing), then ties
    def py_sort_key(r):
        status = r.get("status", "")
        priority = {
            "restatement": 0,
            "missing-on-fy24-pdf": 1,
            "exception": 2,
            "ties-caption-changed": 3,
            "ties-with-sign-inversion": 4,
            "ties-with-rounding": 5,
            "ties": 6,
        }.get(status, 9)
        return (priority, r.get("pdf_section") or "", r.get("pdf_page") or 0, r.get("pdf_label") or "")
    py_records.sort(key=py_sort_key)

    for r_idx, rec in enumerate(py_records, start=py_header_row + 1):
        for c_idx, (_, field) in enumerate(COLUMNS, start=1):
            v = rec.get(field)
            cell = ws_py.cell(row=r_idx, column=c_idx, value=v)
            cell.border = border
            if field == "status":
                status = v
                if status == "restatement":
                    cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                    cell.font = Font(bold=True, color="A52022")
                elif status == "missing-on-fy24-pdf":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.font = Font(bold=True, color="9C6500")
                elif status == "ties-caption-changed":
                    cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
                elif status == "ties":
                    cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                    cell.font = Font(color="2A7A2A")
    for i, w in enumerate(widths, start=1):
        ws_py.column_dimensions[get_column_letter(i)].width = w
    ws_py.freeze_panes = f"A{py_header_row + 1}"
    # PY tab summary at top
    py_total = len(py_records)
    py_ties = sum(1 for r in py_records if r["status"] == "ties")
    py_round = sum(1 for r in py_records if r["status"] in ("ties-with-rounding", "ties-with-sign-inversion"))
    py_caption = sum(1 for r in py_records if r["status"] == "ties-caption-changed")
    py_restate = sum(1 for r in py_records if r["status"] == "restatement")
    py_missing = sum(1 for r in py_records if r["status"] == "missing-on-fy24-pdf")
    ws_py["A3"] = (f"Total: {py_total}  |  ties: {py_ties}  |  ties-with-rounding/sign-inv: {py_round}  |  "
                   f"ties-caption-changed: {py_caption}  |  restatements: {py_restate}  |  missing: {py_missing}")
    ws_py["A3"].font = Font(bold=True)

    # ==== Bridge Ties tab — every PDF face/footnote line ↔ Bridge workbook record ====
    bridge_records = [r for r in all_records if r.get("lane") == "pdf_to_bridge"]
    ws_br = wb.create_sheet("Bridge Ties", 1)
    ws_br["A1"] = "Bridge Ties — every PDF line ↔ Bridge workbook row (Lane 1)"
    ws_br["A1"].font = Font(bold=True, size=13)
    ws_br["A2"] = (
        "'ties' = exact label & number match  |  'ties-with-rounding' = bridge has more precision but rounds to PDF  |  "
        "'ties-caption-changed' = number matches but bridge label differs (e.g., bridge uses 'Total payments' for what FS calls 'Outstanding borrowings')  |  "
        "'exception' = number doesn't tie  |  "
        "'missing-on-bridge' = no bridge row found at all"
    )
    ws_br["A2"].font = Font(italic=True, color="606060")
    ws_br.merge_cells("A2:O2")
    ws_br.row_dimensions[2].height = 36
    br_header_row = 4
    for c, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws_br.cell(row=br_header_row, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Sort: real exceptions first, then ties-caption-changed (worth a review), then ties
    def br_sort_key(r):
        status = r.get("status", "")
        priority = {
            "exception": 0,
            "missing-on-bridge": 1,
            "missing-year-on-bridge": 2,
            "ties-caption-changed": 3,
            "ties-with-rounding": 4,
            "ties-with-sign-inversion": 5,
            "ties": 6,
        }.get(status, 9)
        return (priority, r.get("pdf_section") or "", r.get("pdf_page") or 0,
                r.get("pdf_label") or "", r.get("pdf_year") or "")
    bridge_records.sort(key=br_sort_key)

    for r_idx, rec in enumerate(bridge_records, start=br_header_row + 1):
        for c_idx, (_, field) in enumerate(COLUMNS, start=1):
            v = rec.get(field)
            cell = ws_br.cell(row=r_idx, column=c_idx, value=v)
            cell.border = border
            if field == "status":
                if v == "exception":
                    cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                    cell.font = Font(bold=True, color="A52022")
                elif v in ("missing-on-bridge", "missing-year-on-bridge"):
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.font = Font(bold=True, color="9C6500")
                elif v == "ties-caption-changed":
                    cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
                elif v == "ties-with-rounding":
                    cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
                elif v == "ties":
                    cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                    cell.font = Font(color="2A7A2A")
    for i, w in enumerate(widths, start=1):
        ws_br.column_dimensions[get_column_letter(i)].width = w
    ws_br.freeze_panes = f"A{br_header_row + 1}"

    # Bridge tab summary
    br_total = len(bridge_records)
    br_ties = sum(1 for r in bridge_records if r["status"] == "ties")
    br_round = sum(1 for r in bridge_records if r["status"] == "ties-with-rounding")
    br_cap = sum(1 for r in bridge_records if r["status"] == "ties-caption-changed")
    br_exc = sum(1 for r in bridge_records if r["status"] == "exception")
    br_miss = sum(1 for r in bridge_records if r["status"] in ("missing-on-bridge", "missing-year-on-bridge"))
    ws_br["A3"] = (f"Total: {br_total}  |  ties: {br_ties}  |  ties-with-rounding: {br_round}  |  "
                   f"ties-caption-changed: {br_cap}  |  exceptions: {br_exc}  |  missing: {br_miss}")
    ws_br["A3"].font = Font(bold=True)

    # ==== TB Ties tab — every Bridge BS line ↔ Trial Balance rollup (Lane 2) ====
    tb_records = [r for r in all_records if r.get("lane") == "bridge_to_tb"]
    ws_tb = wb.create_sheet("TB Ties", 1)
    ws_tb["A1"] = "TB Ties — Bridge BS line ↔ Trial Balance rollup (Lane 2)"
    ws_tb["A1"].font = Font(bold=True, size=13)
    ws_tb["A2"] = (
        "'ties' = bridge value matches TB rollup (within tolerance)  |  "
        "'ties-with-rounding' = matches within $5K (subtotal) / $1K (line)  |  "
        "'ties-no-tb-needed' = subtotal (computed) or equity component (ties via SOE)  |  "
        "'exception' = bridge ≠ TB rollup (investigate)  |  "
        "'no-tb-rollup' = bridge line has no TB accounts mapped (and is NOT a subtotal — bridge mapping gap)"
    )
    ws_tb["A2"].font = Font(italic=True, color="606060")
    ws_tb.merge_cells("A2:O2")
    ws_tb.row_dimensions[2].height = 40
    tb_header_row = 4
    for c, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws_tb.cell(row=tb_header_row, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Sort: exceptions first, then no-tb-rollup, then ties (with subtotals last)
    def tb_sort_key(r):
        status = r.get("status", "")
        priority = {
            "exception": 0,
            "no-tb-rollup": 1,
            "ties-with-rounding": 2,
            "ties": 3,
            "ties-no-tb-needed": 4,
        }.get(status, 9)
        return (priority, r.get("pdf_label") or "")
    tb_records.sort(key=tb_sort_key)

    for r_idx, rec in enumerate(tb_records, start=tb_header_row + 1):
        for c_idx, (_, field) in enumerate(COLUMNS, start=1):
            v = rec.get(field)
            cell = ws_tb.cell(row=r_idx, column=c_idx, value=v)
            cell.border = border
            if field == "status":
                if v == "exception":
                    cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                    cell.font = Font(bold=True, color="A52022")
                elif v == "no-tb-rollup":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.font = Font(bold=True, color="9C6500")
                elif v == "ties-no-tb-needed":
                    cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
                    cell.font = Font(color="3B4C68")
                elif v == "ties-with-rounding":
                    cell.fill = PatternFill(start_color="FFFBE5", end_color="FFFBE5", fill_type="solid")
                elif v == "ties":
                    cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                    cell.font = Font(color="2A7A2A")
    for i, w in enumerate(widths, start=1):
        ws_tb.column_dimensions[get_column_letter(i)].width = w
    ws_tb.freeze_panes = f"A{tb_header_row + 1}"

    tb_total = len(tb_records)
    tb_ties = sum(1 for r in tb_records if r["status"] == "ties")
    tb_round = sum(1 for r in tb_records if r["status"] == "ties-with-rounding")
    tb_no_need = sum(1 for r in tb_records if r["status"] == "ties-no-tb-needed")
    tb_exc = sum(1 for r in tb_records if r["status"] == "exception")
    tb_miss = sum(1 for r in tb_records if r["status"] == "no-tb-rollup")
    ws_tb["A3"] = (f"Total: {tb_total}  |  ties: {tb_ties}  |  ties-with-rounding: {tb_round}  |  "
                   f"ties-no-tb-needed: {tb_no_need}  |  exceptions: {tb_exc}  |  no-tb-rollup: {tb_miss}")
    ws_tb["A3"].font = Font(bold=True)

    # ==== Mapping Completeness tab — Lane 7 contextual / "did anything fall off" checks ====
    mc_records = [r for r in all_records if r.get("lane") == "mapping_completeness"]
    ws_mc = wb.create_sheet("Mapping Completeness", 1)
    ws_mc["A1"] = "Mapping Completeness — did every TB account reach the FS, and does it all make sense? (Lane 7)"
    ws_mc["A1"].font = Font(bold=True, size=13)
    ws_mc["A2"] = (
        "'unmapped-account' = TB account with a balance not in the bridge 'TB Mapping' tab (a new account that fell off the FS)  |  "
        "'mapped-to-nothing' = in the mapping but no FS target  |  'completeness-gap' = gross $ not reaching the FS  |  "
        "'stale-mapping-target' = mapping points to a BS line that no longer exists  |  'tb-out-of-balance' / 'bs-does-not-balance' = integrity/identity failed  |  "
        "'ties' = check passed"
    )
    ws_mc["A2"].font = Font(italic=True, color="606060")
    ws_mc.merge_cells("A2:O2")
    ws_mc.row_dimensions[2].height = 40
    mc_header_row = 4
    for c, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws_mc.cell(row=mc_header_row, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    def mc_sort_key(r):
        status = r.get("status", "")
        priority = {
            "unmapped-account": 0,
            "mapped-to-nothing": 1,
            "completeness-gap": 2,
            "tb-out-of-balance": 3,
            "bs-does-not-balance": 4,
            "stale-mapping-target": 5,
            "ties": 6,
        }.get(status, 9)
        return (priority, r.get("pdf_label") or "")
    mc_records.sort(key=mc_sort_key)

    for r_idx, rec in enumerate(mc_records, start=mc_header_row + 1):
        for c_idx, (_, field) in enumerate(COLUMNS, start=1):
            v = rec.get(field)
            cell = ws_mc.cell(row=r_idx, column=c_idx, value=v)
            cell.border = border
            if field == "status":
                if v in ("unmapped-account", "completeness-gap", "tb-out-of-balance", "bs-does-not-balance"):
                    cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                    cell.font = Font(bold=True, color="A52022")
                elif v in ("mapped-to-nothing", "stale-mapping-target"):
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.font = Font(bold=True, color="9C6500")
                elif v == "ties":
                    cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                    cell.font = Font(color="2A7A2A")
    for i, w in enumerate(widths, start=1):
        ws_mc.column_dimensions[get_column_letter(i)].width = w
    ws_mc.freeze_panes = f"A{mc_header_row + 1}"

    mc_unmapped = sum(1 for r in mc_records if r["status"] == "unmapped-account")
    mc_nothing = sum(1 for r in mc_records if r["status"] == "mapped-to-nothing")
    mc_stale = sum(1 for r in mc_records if r["status"] == "stale-mapping-target")
    mc_fail = sum(1 for r in mc_records if r["status"] in ("completeness-gap", "tb-out-of-balance", "bs-does-not-balance"))
    mc_verdict = "ALL CLEAR" if (mc_unmapped + mc_nothing + mc_stale + mc_fail) == 0 else "REVIEW NEEDED"
    ws_mc["A3"] = (f"{mc_verdict}  |  unmapped accounts: {mc_unmapped}  |  mapped-to-nothing: {mc_nothing}  |  "
                   f"stale targets: {mc_stale}  |  integrity/identity failures: {mc_fail}")
    ws_mc["A3"].font = Font(bold=True, color=("2A7A2A" if mc_verdict == "ALL CLEAR" else "A52022"))

    # ==== SOE Rollforward tab — every Lane 5 SOE record ====
    soe_records = [r for r in all_records if r.get("lane") == "soe"]
    ws_soe = wb.create_sheet("SOE Rollforward", 1)
    ws_soe["A1"] = "SOE Rollforward — Members' Equity changes by class (Lane 5)"
    ws_soe["A1"].font = Font(bold=True, size=13)
    ws_soe["A2"] = (
        "Each value on the SOE (per class column × per row) is tied to the Bridge SOE tab. "
        "Cross-foot (xF) verifies Total = sum(SCU + A-1 + A-2 + B + C + D + AccDef + AOCI). "
        "Roll-forward (F) verifies Opening + Activities = Closing for each year."
    )
    ws_soe["A2"].font = Font(italic=True, color="606060")
    ws_soe.merge_cells("A2:O2")
    ws_soe.row_dimensions[2].height = 36
    soe_header_row = 4
    for c, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws_soe.cell(row=soe_header_row, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    def soe_sort_key(r):
        status = r.get("status", "")
        priority = {
            "exception": 0,
            "missing-on-bridge": 1,
            "ties-F": 2,
            "ties-xF": 3,
            "ties-with-rounding": 4,
            "ties": 5,
        }.get(status, 9)
        return (priority, r.get("pdf_year") or "", r.get("pdf_label") or "")
    soe_records.sort(key=soe_sort_key)

    for r_idx, rec in enumerate(soe_records, start=soe_header_row + 1):
        for c_idx, (_, field) in enumerate(COLUMNS, start=1):
            v = rec.get(field)
            cell = ws_soe.cell(row=r_idx, column=c_idx, value=v)
            cell.border = border
            if field == "status":
                if v == "exception":
                    cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                    cell.font = Font(bold=True, color="A52022")
                elif v in ("ties-F", "ties-xF"):
                    cell.fill = PatternFill(start_color="E0F0E0", end_color="E0F0E0", fill_type="solid")
                    cell.font = Font(bold=True, color="2A7A2A")
                elif v == "ties":
                    cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                    cell.font = Font(color="2A7A2A")
                elif v == "ties-with-rounding":
                    cell.fill = PatternFill(start_color="FFFBE5", end_color="FFFBE5", fill_type="solid")
    for i, w in enumerate(widths, start=1):
        ws_soe.column_dimensions[get_column_letter(i)].width = w
    ws_soe.freeze_panes = f"A{soe_header_row + 1}"

    soe_total = len(soe_records)
    soe_ties = sum(1 for r in soe_records if r["status"] == "ties")
    soe_round = sum(1 for r in soe_records if r["status"] == "ties-with-rounding")
    soe_F = sum(1 for r in soe_records if r["status"] == "ties-F")
    soe_xF = sum(1 for r in soe_records if r["status"] == "ties-xF")
    soe_exc = sum(1 for r in soe_records if r["status"] == "exception")
    ws_soe["A3"] = (f"Total: {soe_total}  |  ties: {soe_ties}  |  ties-with-rounding: {soe_round}  |  "
                    f"ties-F (rollforward): {soe_F}  |  ties-xF (cross-foot): {soe_xF}  |  exceptions: {soe_exc}")
    ws_soe["A3"].font = Font(bold=True)

    # ==== Internal Ties tab — every internal cross-reference within the PDF ====
    internal_records = [r for r in all_records if r.get("lane") == "pdf_internal"]
    ws_int = wb.create_sheet("Internal Ties", 1)  # will end up at index after re-ordering
    ws_int["A1"] = "Internal Cross-References — values that should agree between PDF sections"
    ws_int["A1"].font = Font(bold=True, size=13)
    ws_int["A2"] = (
        "Mark vocabulary (mirrors FY24):  /SoCF → ties to Stmt of Cash Flows  |  "
        "/SoSE → ties to Stmt of Members' Equity  |  /PL → ties to Income Stmt  |  "
        "/BS → ties to Balance Sheet  |  /FN N → ties to Footnote N  |  "
        "F → internal subtotal foots"
    )
    ws_int["A2"].font = Font(italic=True, color="606060")
    ws_int.merge_cells("A2:P2")
    ws_int.row_dimensions[2].height = 30

    # Internal Ties tab columns: include the Mark column
    INTERNAL_COLUMNS = [
        ("Lane", "lane"),
        ("PDF Page", "pdf_page"),
        ("Section", "pdf_section"),
        ("Mark", "mark"),
        ("Label", "pdf_label"),
        ("Year", "pdf_year"),
        ("PDF Value", "pdf_value"),
        ("Target Section", "target_section"),
        ("Target Page", "target_page"),
        ("Source Label", "source_label"),
        ("Source Value", "source_value"),
        ("Delta", "delta"),
        ("Tolerance", "tolerance"),
        ("Status", "status"),
        ("Notes", "notes"),
    ]
    int_header_row = 4
    for c, (label, _) in enumerate(INTERNAL_COLUMNS, start=1):
        cell = ws_int.cell(row=int_header_row, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Sort: exceptions first, then ties grouped by mark/section/page
    def internal_sort_key(r):
        status = r.get("status", "")
        priority = {
            "exception": 0,
            "ties": 1,
            "ties-with-rounding": 2,
        }.get(status, 9)
        return (priority, r.get("pdf_section") or "", r.get("pdf_page") or 0,
                r.get("mark") or "", r.get("pdf_label") or "")
    internal_records.sort(key=internal_sort_key)

    for r_idx, rec in enumerate(internal_records, start=int_header_row + 1):
        for c_idx, (_, field) in enumerate(INTERNAL_COLUMNS, start=1):
            v = rec.get(field)
            cell = ws_int.cell(row=r_idx, column=c_idx, value=v)
            cell.border = border
            if field == "status":
                if v == "exception":
                    cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                    cell.font = Font(bold=True, color="A52022")
                elif v == "ties":
                    cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                    cell.font = Font(color="2A7A2A")
                elif v == "ties-with-rounding":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    # Column widths for Internal Ties tab
    int_widths = [12, 8, 8, 10, 40, 6, 14, 10, 8, 35, 14, 10, 10, 14, 40]
    for i, w in enumerate(int_widths, start=1):
        ws_int.column_dimensions[get_column_letter(i)].width = w
    ws_int.freeze_panes = f"A{int_header_row + 1}"

    # Internal Ties tab summary at top
    int_total = len(internal_records)
    int_ties = sum(1 for r in internal_records if r["status"] in ("ties", "ties-with-rounding"))
    int_exc = sum(1 for r in internal_records if r["status"] == "exception")
    # Count by mark
    mark_counts = {}
    for r in internal_records:
        mk = r.get("mark", "?")
        mark_counts[mk] = mark_counts.get(mk, 0) + 1
    marks_summary = ", ".join(f"{m}={c}" for m, c in sorted(mark_counts.items()))
    ws_int["A3"] = f"Total: {int_total}  |  ties: {int_ties}  |  exceptions: {int_exc}  |  By mark: {marks_summary}"
    ws_int["A3"].font = Font(bold=True)

    # Footing tab — segregate all footing/cross-foot checks
    footing_records = [r for r in all_records
                       if r.get("lane") == "footing"
                       or r.get("status") in ("ties-F", "ties-xF")
                       or (r.get("status") == "exception" and r.get("pdf_label", "").endswith("(F)"))
                       or (r.get("status") == "exception" and r.get("pdf_label", "").endswith("(xF)"))]
    ws_foot = wb.create_sheet("Footing", 1)  # insert as second tab (right after Summary)
    ws_foot["A1"] = "Footing & Cross-Foot Checks"
    ws_foot["A1"].font = Font(bold=True, size=13)
    ws_foot["A2"] = (
        "Every subtotal verified by summing components. "
        "'ties-F' = footed; 'ties-xF' = cross-foot recalc; 'exception' = math fails."
    )
    ws_foot["A2"].font = Font(italic=True, color="606060")
    # Header row at row 4
    foot_header_row = 4
    for c, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws_foot.cell(row=foot_header_row, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Sort by section, then page, then label
    footing_records.sort(key=lambda r: (
        r.get("pdf_section") or "",
        r.get("pdf_page") or 0,
        r.get("pdf_label") or "",
        r.get("pdf_year") or "",
    ))
    for r_idx, rec in enumerate(footing_records, start=foot_header_row + 1):
        for c_idx, (_, field) in enumerate(COLUMNS, start=1):
            v = rec.get(field)
            cell = ws_foot.cell(row=r_idx, column=c_idx, value=v)
            cell.border = border
            # Color status cell
            if field == "status":
                if v in ("ties-F", "ties-xF"):
                    cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                    cell.font = Font(bold=True, color="2A7A2A")
                elif v == "exception":
                    cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                    cell.font = Font(bold=True, color="A52022")
    for i, w in enumerate(widths, start=1):
        ws_foot.column_dimensions[get_column_letter(i)].width = w
    ws_foot.freeze_panes = f"A{foot_header_row + 1}"
    # Footing-tab summary at top
    f_total = len(footing_records)
    f_ties_F = sum(1 for r in footing_records if r["status"] == "ties-F")
    f_ties_xF = sum(1 for r in footing_records if r["status"] == "ties-xF")
    f_exc = sum(1 for r in footing_records if r["status"] == "exception")
    ws_foot["A3"] = f"Total checks: {f_total}  |  ties-F: {f_ties_F}  |  ties-xF: {f_ties_xF}  |  exceptions: {f_exc}"
    ws_foot["A3"].font = Font(bold=True)

    # ==== PBC Ties tab — Lane 8: bridge footnote ↔ supporting PBC ↔ GL ====
    pbc_records = [r for r in all_records if r.get("lane") == "pbc_to_bridge"]
    if pbc_records:
        ws_pt = wb.create_sheet("PBC Ties", 1)
        ws_pt["A1"] = "PBC Ties — bridge footnote ↔ supporting PBC workpaper ↔ GL (Lane 8)"
        ws_pt["A1"].font = Font(bold=True, size=13)
        ws_pt["A2"] = (
            "Closes the source-to-disclosure gap: ties each mapped footnote back to the PBC that "
            "supports it.  '(subledger vs GL)' = the company's rec ties its own ledger  |  "
            "'(bridge vs PBC)' = the disclosed footnote ties the supporting workpaper  |  "
            "'DIAGNOSTIC: offsetting gross/accum' = net ties but gross & accumulated amortization "
            "each diverge by the same amount — the fully-amortized/disposed-asset gotcha "
            "(see source-of-truth-hierarchy.md). Bridge values in $K."
        )
        ws_pt["A2"].font = Font(italic=True, color="606060")
        ws_pt.merge_cells("A2:O2")
        ws_pt.row_dimensions[2].height = 48
        pt_header_row = 4
        for c, (label, _) in enumerate(COLUMNS, start=1):
            cell = ws_pt.cell(row=pt_header_row, column=c, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        def pt_sort_key(r):
            status = r.get("status", "")
            # Diagnostics + exceptions first
            is_diag = "DIAGNOSTIC" in (r.get("pdf_label") or "")
            priority = {"exception": 0, "missing": 1,
                        "ties-with-rounding": 3, "ties": 4}.get(status, 2)
            if is_diag:
                priority = -1
            return (priority, r.get("pdf_section") or "", r.get("pdf_label") or "")
        pbc_records.sort(key=pt_sort_key)

        for r_idx, rec in enumerate(pbc_records, start=pt_header_row + 1):
            for c_idx, (_, field) in enumerate(COLUMNS, start=1):
                v = rec.get(field)
                if field in ("pdf_value", "source_value", "delta", "tolerance") and v is not None:
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                cell = ws_pt.cell(row=r_idx, column=c_idx, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(field in ("notes", "pdf_label")))
                if field == "status":
                    if v == "exception":
                        cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                        cell.font = Font(bold=True, color="A52022")
                    elif v == "ties-with-rounding":
                        cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
                    elif v == "ties":
                        cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                        cell.font = Font(color="2A7A2A")
        for i, w in enumerate(widths, start=1):
            ws_pt.column_dimensions[get_column_letter(i)].width = w
        ws_pt.freeze_panes = f"A{pt_header_row + 1}"

        pt_total = len(pbc_records)
        pt_ties = sum(1 for r in pbc_records if r["status"] in ("ties", "ties-with-rounding"))
        pt_exc = sum(1 for r in pbc_records if r["status"] == "exception")
        ws_pt["A3"] = (f"Total: {pt_total}  |  ties: {pt_ties}  |  exceptions: {pt_exc}"
                       + ("  ← includes the intangibles gross/accum overstatement finding" if pt_exc else ""))
        ws_pt["A3"].font = Font(bold=True, color=("A52022" if pt_exc else "2A7A2A"))

    # ==== Flux Review tab — Lane 10: YoY movements recomputed on the FINAL FS ====
    flux_records = [r for r in all_records if r.get("lane") == "flux"]
    if flux_records:
        ws_fx = wb.create_sheet("Flux Review", 1)
        ws_fx["A1"] = "Flux Review — YoY movements recomputed on the FINAL FS (2025 vs 2024)"
        ws_fx["A1"].font = Font(bold=True, size=13)
        ws_fx["A2"] = (
            "Re-baselines the company's flux PBC to the final numbers — the PBC is drafted before the "
            "revision rounds, so it goes stale and will NOT tie.  'mover-needs-explanation' = material "
            "move with no carried comment/driver — draft one  |  'mover-explained' = a related prior "
            "comment (directional, from the stale PBC) or a context driver was found  |  "
            "'mover-subtotal' = explained by its components  |  'immaterial' = below threshold. "
            "Carried comments are CONTEXT, not a tie. Values in $K."
        )
        ws_fx["A2"].font = Font(italic=True, color="606060")
        ws_fx.merge_cells("A2:O2")
        ws_fx.row_dimensions[2].height = 56
        fx_header_row = 4
        FLUX_COLUMNS = [
            ("Statement", "pdf_section"), ("Line", "pdf_label"), ("Period", "pdf_year"),
            ("Current ($K)", "pdf_value"), ("Prior ($K)", "source_value"),
            ("$ Var", "delta"), ("Threshold", "tolerance"), ("Status", "status"),
            ("Comment / suggested driver", "notes"),
        ]
        for c, (label, _) in enumerate(FLUX_COLUMNS, start=1):
            cell = ws_fx.cell(row=fx_header_row, column=c, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        def fx_sort_key(r):
            pri = {"mover-needs-explanation": 0, "mover-explained": 1,
                   "mover-subtotal": 2, "immaterial": 3}.get(r.get("status", ""), 4)
            return (pri, r.get("pdf_section") or "", -abs(r.get("delta") or 0))
        flux_records.sort(key=fx_sort_key)

        for r_idx, rec in enumerate(flux_records, start=fx_header_row + 1):
            for c_idx, (_, field) in enumerate(FLUX_COLUMNS, start=1):
                v = rec.get(field)
                if field in ("pdf_value", "source_value", "delta", "tolerance") and v is not None:
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                cell = ws_fx.cell(row=r_idx, column=c_idx, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(field in ("notes", "pdf_label")))
                if field == "status":
                    if v == "mover-needs-explanation":
                        cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                        cell.font = Font(bold=True, color="A52022")
                    elif v == "mover-explained":
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        cell.font = Font(color="9C6500")
                    elif v == "mover-subtotal":
                        cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
                    elif v == "immaterial":
                        cell.fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
                        cell.font = Font(color="808080")
        fx_widths = [16, 40, 12, 13, 13, 11, 11, 24, 70]
        for i, w in enumerate(fx_widths, start=1):
            ws_fx.column_dimensions[get_column_letter(i)].width = w
        ws_fx.freeze_panes = f"A{fx_header_row + 1}"

        fx_need = sum(1 for r in flux_records if r["status"] == "mover-needs-explanation")
        fx_exp = sum(1 for r in flux_records if r["status"] == "mover-explained")
        fx_sub = sum(1 for r in flux_records if r["status"] == "mover-subtotal")
        ws_fx["A3"] = (f"FS lines: {len(flux_records)}  |  need explanation: {fx_need}  |  "
                       f"carried/context: {fx_exp}  |  subtotal movers: {fx_sub}")
        ws_fx["A3"].font = Font(bold=True, color=("A52022" if fx_need else "2A7A2A"))

    # ==== State Tax tab — engine Module F: blended state rate derivation ====
    sx_records = [r for r in all_records if r.get("lane") == "state_tax"]
    if sx_records:
        ws_sx = wb.create_sheet("State Tax", 1)
        ws_sx["A1"] = "State Tax — blended state rate derivation (apportionment × rate → ETR)"
        ws_sx["A1"].font = Font(bold=True, size=13)
        ws_sx["A2"] = (
            "Validates the blended state rate that feeds the provision: every state's ETR = "
            "apportionment factor × state statutory rate, the per-state ETRs sum to the blended state "
            "rate, and federal + state × (1 − fed) ties the rate-rec blended rate. The per-state "
            "apportionment factor from raw sales (sourcing / throwback) stays judgment. Rates shown."
        )
        ws_sx["A2"].font = Font(italic=True, color="606060")
        ws_sx.merge_cells("A2:O2")
        ws_sx.row_dimensions[2].height = 44
        sx_header_row = 4
        for c, (label, _) in enumerate(COLUMNS, start=1):
            cell = ws_sx.cell(row=sx_header_row, column=c, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        def sx_sort_key(r):
            return ({"exception": 0}.get(r.get("status", ""), 1), r.get("pdf_label") or "")
        sx_records.sort(key=sx_sort_key)

        for r_idx, rec in enumerate(sx_records, start=sx_header_row + 1):
            for c_idx, (_, field) in enumerate(COLUMNS, start=1):
                v = rec.get(field)
                if field in ("pdf_value", "source_value", "delta", "tolerance") and v is not None:
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                cell = ws_sx.cell(row=r_idx, column=c_idx, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(field in ("notes", "pdf_label")))
                if field == "status":
                    if v == "exception":
                        cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                        cell.font = Font(bold=True, color="A52022")
                    elif v in ("ties", "ties-with-rounding"):
                        cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                        cell.font = Font(color="2A7A2A")
        for i, w in enumerate(widths, start=1):
            ws_sx.column_dimensions[get_column_letter(i)].width = w
        ws_sx.freeze_panes = f"A{sx_header_row + 1}"

        sx_exc = sum(1 for r in sx_records if r["status"] == "exception")
        ws_sx["A3"] = (f"State-rate checks: {len(sx_records) - sx_exc} tie / {sx_exc} exception")
        ws_sx["A3"].font = Font(bold=True, color=("A52022" if sx_exc else "2A7A2A"))

    # ==== Rate Rec tab — engine Module E (capstone): statutory->effective assembly ====
    rr_records = [r for r in all_records if r.get("lane") == "rate_rec"]
    if rr_records:
        ws_rr = wb.create_sheet("Rate Rec", 1)
        ws_rr["A1"] = "Rate Rec — statutory→effective reconciliation, independently assembled (capstone)"
        ws_rr["A1"].font = Font(bold=True, size=13)
        ws_rr["A2"] = (
            "Assembles the whole rate reconciliation from its parts and confirms it foots to the total "
            "provision and ties the FS.  '[independent]' rows (statutory + perms) are recomputed from "
            "source ('ties-recomputed'); '[workpaper]' rows (state, true-ups, DTA adj, change in VA, "
            "rate change, foreign) are taken from the preparer's other schedules ('ties-rate-only') and "
            "are the next tier to recompute.  The footing + FS-tie rows close the loop. Values in $1."
        )
        ws_rr["A2"].font = Font(italic=True, color="606060")
        ws_rr.merge_cells("A2:O2")
        ws_rr.row_dimensions[2].height = 52
        rr_header_row = 4
        for c, (label, _) in enumerate(COLUMNS, start=1):
            cell = ws_rr.cell(row=rr_header_row, column=c, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        def rr_sort_key(r):
            lbl = r.get("pdf_label") or ""
            # keep the assembly order: statutory, perms, others, then footing/FS/ETR summary rows
            grp = (0 if "statutory rate" in lbl else 1 if lbl.startswith("Perm:")
                   else 2 if lbl.startswith("Other:") else 3)
            return ({"exception": 0}.get(r.get("status", ""), 1), grp, lbl)
        rr_records.sort(key=rr_sort_key)

        for r_idx, rec in enumerate(rr_records, start=rr_header_row + 1):
            for c_idx, (_, field) in enumerate(COLUMNS, start=1):
                v = rec.get(field)
                if field in ("pdf_value", "source_value", "delta", "tolerance") and v is not None:
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                cell = ws_rr.cell(row=r_idx, column=c_idx, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(field in ("notes", "pdf_label")))
                if field == "status":
                    if v == "exception":
                        cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                        cell.font = Font(bold=True, color="A52022")
                    elif v == "ties-rate-only":
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        cell.font = Font(color="9C6500")
                    elif v == "ties-recomputed":
                        cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                        cell.font = Font(bold=True, color="2A7A2A")
                    elif v in ("ties", "ties-with-rounding"):
                        cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                        cell.font = Font(color="2A7A2A")
        for i, w in enumerate(widths, start=1):
            ws_rr.column_dimensions[get_column_letter(i)].width = w
        ws_rr.freeze_panes = f"A{rr_header_row + 1}"

        rr_indep = sum(1 for r in rr_records if r["status"] == "ties-recomputed")
        rr_wp = sum(1 for r in rr_records if r["status"] == "ties-rate-only")
        rr_exc = sum(1 for r in rr_records if r["status"] == "exception")
        ws_rr["A3"] = (f"Independent (recomputed from source): {rr_indep}  |  workpaper-sourced: {rr_wp}  |  "
                       f"exceptions: {rr_exc}  |  rate rec foots + ties FS")
        ws_rr["A3"].font = Font(bold=True, color=("A52022" if rr_exc else "2A7A2A"))

    # ==== Current & NOL tab — engine Module D: taxable-income build + NOL trend ====
    cn_records = [r for r in all_records if r.get("lane") == "current_nol"]
    if cn_records:
        ws_cn = wb.create_sheet("Current & NOL", 1)
        ws_cn["A1"] = "Current & NOL — taxable-income build + federal NOL position + NOL DTA trend"
        ws_cn["A1"].font = Font(bold=True, size=13)
        ws_cn["A2"] = (
            "Recomputes the federal taxable-income build (pretax + permanent + temporary differences "
            "= taxable income) and confirms a loss year carries no federal cash tax.  'nol-trend' = "
            "NOL DTA across FY23/24/25 (a loss company's NOL normally accretes); 'nol-decrease' flags "
            "a drop to explain (utilization / expiry / write-off). The per-jurisdiction current split "
            "and the NOL vintage roll (RTP / apportionment) are next-tier. Values in $1."
        )
        ws_cn["A2"].font = Font(italic=True, color="606060")
        ws_cn.merge_cells("A2:O2")
        ws_cn.row_dimensions[2].height = 52
        cn_header_row = 4
        for c, (label, _) in enumerate(COLUMNS, start=1):
            cell = ws_cn.cell(row=cn_header_row, column=c, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        def cn_sort_key(r):
            return ({"exception": 0, "nol-decrease": 1, "ties": 3, "nol-trend": 4}
                    .get(r.get("status", ""), 2), r.get("pdf_label") or "")
        cn_records.sort(key=cn_sort_key)

        for r_idx, rec in enumerate(cn_records, start=cn_header_row + 1):
            for c_idx, (_, field) in enumerate(COLUMNS, start=1):
                v = rec.get(field)
                if field in ("pdf_value", "source_value", "delta", "tolerance") and v is not None:
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                cell = ws_cn.cell(row=r_idx, column=c_idx, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(field in ("notes", "pdf_label")))
                if field == "status":
                    if v == "exception":
                        cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                        cell.font = Font(bold=True, color="A52022")
                    elif v == "nol-decrease":
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        cell.font = Font(color="9C6500")
                    elif v == "nol-trend":
                        cell.fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
                        cell.font = Font(color="808080")
                    elif v == "ties":
                        cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                        cell.font = Font(color="2A7A2A")
        for i, w in enumerate(widths, start=1):
            ws_cn.column_dimensions[get_column_letter(i)].width = w
        ws_cn.freeze_panes = f"A{cn_header_row + 1}"

        cn_tie = sum(1 for r in cn_records if r["status"] == "ties")
        cn_exc = sum(1 for r in cn_records if r["status"] == "exception")
        cn_dec = sum(1 for r in cn_records if r["status"] == "nol-decrease")
        ws_cn["A3"] = (f"Build checks: {cn_tie} tie / {cn_exc} exception  |  NOL trend flags: {cn_dec}")
        ws_cn["A3"].font = Font(bold=True, color=("A52022" if (cn_exc or cn_dec) else "2A7A2A"))

    # ==== Deferred Tax tab — engine Module C: deferred FS tie + cross-year trend ====
    dt_records = [r for r in all_records if r.get("lane") == "deferred_tax"]
    if dt_records:
        ws_dt = wb.create_sheet("Deferred Tax", 1)
        ws_dt["A1"] = "Deferred Tax — provision DTA/DTL grouping ↔ FS deferred table + 3-year trend"
        ws_dt["A1"].font = Font(bold=True, size=13)
        ws_dt["A2"] = (
            "Reproduces the FN-07 deferred footnote from the provision workpaper (NOL, §163j, R&E, "
            "intangibles, cap commissions, ROU, …) and foots DTA + DTL − VA = net.  The trend rows "
            "track each temporary difference across FY23/24/25 — 'deferred-dropped' = a balance that "
            "was present then zeroed (confirm it was genuinely released, not a dropped roll-forward "
            "line); 'deferred-new' = newly appearing. Tie rows in $K; trend rows in $1."
        )
        ws_dt["A2"].font = Font(italic=True, color="606060")
        ws_dt.merge_cells("A2:O2")
        ws_dt.row_dimensions[2].height = 52
        dt_header_row = 4
        for c, (label, _) in enumerate(COLUMNS, start=1):
            cell = ws_dt.cell(row=dt_header_row, column=c, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        def dt_sort_key(r):
            pri = {"exception": 0, "missing": 1, "deferred-dropped": 2, "deferred-new": 3,
                   "ties-with-rounding": 5, "ties": 6, "deferred-trend": 7}.get(r.get("status", ""), 4)
            return (pri, r.get("pdf_label") or "")
        dt_records.sort(key=dt_sort_key)

        for r_idx, rec in enumerate(dt_records, start=dt_header_row + 1):
            for c_idx, (_, field) in enumerate(COLUMNS, start=1):
                v = rec.get(field)
                if field in ("pdf_value", "source_value", "delta", "tolerance") and v is not None:
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                cell = ws_dt.cell(row=r_idx, column=c_idx, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(field in ("notes", "pdf_label")))
                if field == "status":
                    if v in ("exception", "missing"):
                        cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                        cell.font = Font(bold=True, color="A52022")
                    elif v in ("deferred-dropped", "deferred-new"):
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        cell.font = Font(color="9C6500")
                    elif v == "deferred-trend":
                        cell.fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
                        cell.font = Font(color="808080")
                    elif v in ("ties", "ties-with-rounding"):
                        cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                        cell.font = Font(color="2A7A2A")
        for i, w in enumerate(widths, start=1):
            ws_dt.column_dimensions[get_column_letter(i)].width = w
        ws_dt.freeze_panes = f"A{dt_header_row + 1}"

        dt_tie = sum(1 for r in dt_records if r["status"] in ("ties", "ties-with-rounding"))
        dt_exc = sum(1 for r in dt_records if r["status"] in ("exception", "missing"))
        dt_drop = sum(1 for r in dt_records if r["status"] == "deferred-dropped")
        dt_new = sum(1 for r in dt_records if r["status"] == "deferred-new")
        ws_dt["A3"] = (f"FS-tie: {dt_tie} tie / {dt_exc} exception  |  trend flags: {dt_drop} dropped, "
                       f"{dt_new} new (confirm)")
        ws_dt["A3"].font = Font(bold=True, color=("A52022" if dt_exc else "2A7A2A"))

    # ==== Provision Recompute tab — engine Module B: independent perm recompute ====
    rc_records = [r for r in all_records if r.get("lane") == "tax_recompute"]
    if rc_records:
        ws_rc = wb.create_sheet("Provision Recompute", 1)
        ws_rc["A1"] = "Provision Recompute — permanent differences recomputed independently vs preparer"
        ws_rc["A1"].font = Font(bold=True, size=13)
        ws_rc["A2"] = (
            "The engine recomputes each perm from source and red-boxes drift from the preparer's file "
            "(complements the Tax Provision tab, which ties preparer↔FS).  "
            "'ties-recomputed' = book pulled straight from the GL and the recomputed tax effect ties — "
            "fully independent, no workpaper trust  |  'ties-rate-only' = the tax-effecting rate is "
            "verified but the book amount couldn't be GL-sourced yet (an aggregation rule or sub-schedule "
            "is needed — see Notes)  |  'exception' = does not reconcile. Values in $1."
        )
        ws_rc["A2"].font = Font(italic=True, color="606060")
        ws_rc.merge_cells("A2:O2")
        ws_rc.row_dimensions[2].height = 52
        rc_header_row = 4
        for c, (label, _) in enumerate(COLUMNS, start=1):
            cell = ws_rc.cell(row=rc_header_row, column=c, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        def rc_sort_key(r):
            return ({"exception": 0, "ties-rate-only": 1, "ties-recomputed": 2}
                    .get(r.get("status", ""), 3), r.get("pdf_label") or "")
        rc_records.sort(key=rc_sort_key)

        for r_idx, rec in enumerate(rc_records, start=rc_header_row + 1):
            for c_idx, (_, field) in enumerate(COLUMNS, start=1):
                v = rec.get(field)
                if field in ("pdf_value", "source_value", "delta", "tolerance") and v is not None:
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                cell = ws_rc.cell(row=r_idx, column=c_idx, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(field in ("notes", "pdf_label")))
                if field == "status":
                    if v == "exception":
                        cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                        cell.font = Font(bold=True, color="A52022")
                    elif v == "ties-rate-only":
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        cell.font = Font(color="9C6500")
                    elif v == "ties-recomputed":
                        cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                        cell.font = Font(bold=True, color="2A7A2A")
        for i, w in enumerate(widths, start=1):
            ws_rc.column_dimensions[get_column_letter(i)].width = w
        ws_rc.freeze_panes = f"A{rc_header_row + 1}"

        rc_indep = sum(1 for r in rc_records if r["status"] == "ties-recomputed")
        rc_rate = sum(1 for r in rc_records if r["status"] == "ties-rate-only")
        rc_exc = sum(1 for r in rc_records if r["status"] == "exception")
        ws_rc["A3"] = (f"Perms: {len(rc_records)}  |  GL-independent & tie: {rc_indep}  |  "
                       f"rate-verified (book not yet GL-sourced): {rc_rate}  |  exceptions: {rc_exc}")
        ws_rc["A3"].font = Font(bold=True, color=("A52022" if rc_exc else "2A7A2A"))

    # ==== Tax Provision tab — Lane 9: FN-07 <-> provision workpaper ====
    tax_records = [r for r in all_records if r.get("lane") == "tax_provision"]
    if tax_records:
        ws_tx = wb.create_sheet("Tax Provision", 1)
        ws_tx["A1"] = "Tax Provision — FN-07 ↔ provision workpaper (rate rec, deferreds, book pretax)"
        ws_tx["A1"].font = Font(bold=True, size=13)
        ws_tx["A2"] = (
            "Re-performs the provision checks an auditor runs: statutory tax recomputes from "
            "pretax × federal rate; the rate reconciliation foots to the total provision; total "
            "provision ties the FS income tax expense; the pretax the provision is built on equals "
            "the FS pretax (the IS↔FN tax-reconciliation error class); and the deferred summary "
            "ties + foots (DTA + DTL − VA = net). All in $K."
        )
        ws_tx["A2"].font = Font(italic=True, color="606060")
        ws_tx.merge_cells("A2:O2")
        ws_tx.row_dimensions[2].height = 48
        tx_header_row = 4
        for c, (label, _) in enumerate(COLUMNS, start=1):
            cell = ws_tx.cell(row=tx_header_row, column=c, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        def tx_sort_key(r):
            return ({"exception": 0, "missing": 1, "ties-with-rounding": 3, "ties": 4}
                    .get(r.get("status", ""), 2), r.get("pdf_label") or "")
        tax_records.sort(key=tx_sort_key)

        for r_idx, rec in enumerate(tax_records, start=tx_header_row + 1):
            for c_idx, (_, field) in enumerate(COLUMNS, start=1):
                v = rec.get(field)
                if field in ("pdf_value", "source_value", "delta", "tolerance") and v is not None:
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                cell = ws_tx.cell(row=r_idx, column=c_idx, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(field in ("notes", "pdf_label")))
                if field == "status":
                    if v == "exception":
                        cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                        cell.font = Font(bold=True, color="A52022")
                    elif v == "missing":
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        cell.font = Font(bold=True, color="9C6500")
                    elif v == "ties-with-rounding":
                        cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
                    elif v == "ties":
                        cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                        cell.font = Font(color="2A7A2A")
        for i, w in enumerate(widths, start=1):
            ws_tx.column_dimensions[get_column_letter(i)].width = w
        ws_tx.freeze_panes = f"A{tx_header_row + 1}"

        tx_ties = sum(1 for r in tax_records if r["status"] in ("ties", "ties-with-rounding"))
        tx_exc = sum(1 for r in tax_records if r["status"] == "exception")
        tx_miss = sum(1 for r in tax_records if r["status"] == "missing")
        ws_tx["A3"] = (f"Total: {len(tax_records)}  |  ties: {tx_ties}  |  exceptions: {tx_exc}"
                       f"  |  missing: {tx_miss}")
        ws_tx["A3"].font = Font(bold=True, color=("A52022" if tx_exc else "2A7A2A"))

    # ==== PBC Register tab — completeness + staleness of the FS-build PBC set ====
    if args.pbc_index and Path(args.pbc_index).exists():
        pbc_recs = json.loads(Path(args.pbc_index).read_text(encoding="utf-8"))
        fs_build = [r for r in pbc_recs if r.get("fs_relevant")]

        # Cutoff for staleness (prep version older than this → re-confirm).
        try:
            cy, cm, cd = (int(x) for x in args.bridge_asof.split("-"))
            cutoff = (cy, cm, cd)
        except Exception:
            cutoff = (2026, 5, 1)

        PBC_COLUMNS = [
            ("FS Area", "fs_area"),
            ("Footnote", "footnote"),
            ("Bridge Tab", "bridge_tab"),
            ("Category", "category"),
            ("Audit Area", "area"),
            ("Filename", "filename"),
            ("Version / As-of", "version"),
            ("Status", "_status"),
            ("Flag", "_flag"),
        ]

        # Build register rows from the FS-build subset.
        register_rows = []
        for r in fs_build:
            prep = parse_prep_date(r.get("version"))
            flag = ""
            if r.get("category") == "flux":
                flag = "possibly-stale (flux PBC drafted pre-revision; re-baseline to final FS)"
            elif prep and prep < cutoff:
                flag = f"possibly-stale (prep {r.get('version')} predates final bridge {args.bridge_asof})"
            elif r.get("category") == "fs-build-unmapped":
                flag = "unmapped — confirm FS area / bridge tab in pbc-mapping.md"
            register_rows.append({**r, "_status": "received", "_flag": flag})

        # Expected-but-missing categories (PBC that should be on hand but isn't present).
        present_cats = {r.get("category") for r in fs_build}
        for cat, label in EXPECTED_FS_BUILD_CATEGORIES.items():
            if cat not in present_cats:
                register_rows.append({
                    "fs_area": label, "footnote": None, "bridge_tab": None,
                    "category": cat, "area": None, "filename": "(none received)",
                    "version": None, "_status": "MISSING", "_flag": "expected FS-build PBC not present",
                })

        ws_pbc = wb.create_sheet("PBC Register", 1)
        ws_pbc["A1"] = "PBC Register — FS-build source workpapers: present? / which FS area / version / stale?"
        ws_pbc["A1"].font = Font(bold=True, size=13)
        ws_pbc["A2"] = (
            "Scope: the FS-build subset only (the PBCs that feed the bridge/footnotes). The ~700 "
            "audit-evidence files (bank statements, invoices, samples, walkthroughs, IT) are indexed "
            "but excluded here.  'received' = file present on disk  |  'MISSING' = expected category "
            "with no file  |  'possibly-stale' = prep version predates the final bridge — re-confirm "
            "(the flux PBC is the canonical stale case)."
        )
        ws_pbc["A2"].font = Font(italic=True, color="606060")
        ws_pbc.merge_cells("A2:I2")
        ws_pbc.row_dimensions[2].height = 44
        pbc_header_row = 4
        for c, (label, _) in enumerate(PBC_COLUMNS, start=1):
            cell = ws_pbc.cell(row=pbc_header_row, column=c, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        def pbc_sort_key(r):
            status_pri = {"MISSING": 0, "received": 1}.get(r["_status"], 2)
            stale_pri = 0 if r["_flag"] else 1
            return (status_pri, stale_pri, r.get("fs_area") or "zzz",
                    r.get("footnote") or "", r.get("filename") or "")
        register_rows.sort(key=pbc_sort_key)

        for r_idx, rec in enumerate(register_rows, start=pbc_header_row + 1):
            for c_idx, (_, field) in enumerate(PBC_COLUMNS, start=1):
                v = rec.get(field)
                cell = ws_pbc.cell(row=r_idx, column=c_idx, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(field in ("_flag", "filename")))
                if field == "_status":
                    if v == "MISSING":
                        cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                        cell.font = Font(bold=True, color="A52022")
                    elif v == "received":
                        cell.fill = PatternFill(start_color="DEFADE", end_color="DEFADE", fill_type="solid")
                        cell.font = Font(color="2A7A2A")
                elif field == "_flag" and v:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.font = Font(color="9C6500")
        pbc_widths = [30, 12, 22, 22, 24, 50, 16, 12, 48]
        for i, w in enumerate(pbc_widths, start=1):
            ws_pbc.column_dimensions[get_column_letter(i)].width = w
        ws_pbc.freeze_panes = f"A{pbc_header_row + 1}"

        n_recv = sum(1 for r in register_rows if r["_status"] == "received")
        n_missing = sum(1 for r in register_rows if r["_status"] == "MISSING")
        n_stale = sum(1 for r in register_rows if r["_flag"].startswith("possibly-stale"))
        n_unmapped = sum(1 for r in register_rows if r.get("category") == "fs-build-unmapped")
        verdict = "ALL EXPECTED PBCs RECEIVED" if n_missing == 0 else f"{n_missing} EXPECTED PBC(s) MISSING"
        ws_pbc["A3"] = (f"{verdict}  |  FS-build files received: {n_recv}  |  missing categories: {n_missing}  |  "
                        f"possibly-stale: {n_stale}  |  unmapped: {n_unmapped}")
        ws_pbc["A3"].font = Font(bold=True, color=("2A7A2A" if n_missing == 0 else "A52022"))
        print(f"  PBC Register: {n_recv} FS-build received, {n_missing} missing, {n_stale} possibly-stale")

    # Summary tab
    ws_sum = wb.create_sheet("Summary", 0)  # insert as first tab
    ws_sum["A1"] = "Tie-out Exceptions Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A3"] = "Status"
    ws_sum["B3"] = "Count"
    ws_sum["A3"].font = Font(bold=True)
    ws_sum["B3"].font = Font(bold=True)
    r = 4
    for status in sorted(summary.keys(), key=lambda s: (s not in TIE_STATUSES, s)):
        ws_sum[f"A{r}"] = status
        ws_sum[f"B{r}"] = summary[status]
        if status not in TIE_STATUSES:
            ws_sum[f"A{r}"].font = Font(bold=True, color="A52022")
        r += 1
    ws_sum[f"A{r+1}"] = f"Total records: {len(all_records)}"
    ws_sum[f"A{r+2}"] = f"Total exceptions: {len(exceptions)}"
    ws_sum[f"A{r+2}"].font = Font(bold=True)
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 10

    # Lane summary
    lane_sum = {}
    for r in all_records:
        key = r["lane"]
        lane_sum.setdefault(key, {"total": 0, "ties": 0, "exceptions": 0})
        lane_sum[key]["total"] += 1
        if r["status"] in TIE_STATUSES or r["status"] in ANALYTICAL_STATUSES:
            lane_sum[key]["ties"] += 1
        else:
            lane_sum[key]["exceptions"] += 1
    r_idx = 12
    ws_sum[f"A{r_idx}"] = "Lane"
    ws_sum[f"B{r_idx}"] = "Total"
    ws_sum[f"C{r_idx}"] = "Ties"
    ws_sum[f"D{r_idx}"] = "Exceptions"
    for col in "ABCD":
        ws_sum[f"{col}{r_idx}"].font = Font(bold=True)
    r_idx += 1
    for lane, stats in sorted(lane_sum.items(), key=lambda x: lane_order.get(x[0], 99)):
        ws_sum[f"A{r_idx}"] = lane
        ws_sum[f"B{r_idx}"] = stats["total"]
        ws_sum[f"C{r_idx}"] = stats["ties"]
        ws_sum[f"D{r_idx}"] = stats["exceptions"]
        r_idx += 1
    ws_sum.column_dimensions["C"].width = 10
    ws_sum.column_dimensions["D"].width = 12

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Wrote {args.out}")
    print(f"  Exceptions sheet: {len(exceptions)} rows")
    print(f"  Summary sheet: status + lane breakdowns")
    if args.include_all:
        print(f"  All Records sheet: {len(all_records)} rows")


if __name__ == "__main__":
    main()
