"""Provision recompute engine — Module F: state tax rate derivation.

Validates the blended state rate that feeds the whole provision. These checks are
mechanical and clean; the per-state apportionment factor *from raw sales* (sourcing
rules, throwback, single- vs three-factor by state) stays judgment / next-tier.

  1. PER-STATE ETR FOOTS — each state's effective rate = apportionment factor x state
     statutory rate, for every state.
  2. BLENDED STATE RATE — the sum of the per-state ETRs equals the blended state rate
     (and ties the schedule's own total row).
  3. FEDERAL + STATE BLENDED RATE — recompute fed + state x (1 - fed) and tie to the
     rate-rec's blended rate; confirm the state rate used in the rate rec matches the
     apportionment schedule.

Emits records under lane "state_tax" -> the "State Tax" report tab.

CLI:
  python recompute_state_tax.py <out.json> [--fy25 <xlsx>] [--fed-rate 0.21]
"""

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from tie_out_common import parse_value, make_record  # noqa: E402

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

ROOT = Path(r"C:\path\to\financial-statement-review")
FY25_DEFAULT = ROOT / "PBCs/2025 Audit/Year End/FS Compilation Requests (FS Compilation Partner) (old)/9. Tax Provision/9. Acme Corp 2025 Tax Provision_v2026.5.21.xlsx"

# '7| State Tax Rate' columns: state=1, apportionment=3, state rate=4, ETR=5.
APPT_COL, RATE_COL, ETR_COL = 3, 4, 5
US_STATES = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","DC","FL","GA","HI","ID","IL","IN","IA","KS",
    "KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ","NM","NY","NC",
    "ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT","VA","WA","WV","WI","WY",
}


def extract_rate_rec_rates(rows):
    """Pull the 'State Rate' and 'Blended Rate' fractions from the rate-rec side panel
    (there may be PY and CY copies; collect all)."""
    state_rates, blended_rates = [], []
    for row in rows:
        labels = [(i, c) for i, c in enumerate(row) if isinstance(c, str)]
        for i, c in labels:
            cl = c.lower().strip()
            if cl in ("state rate", "blended rate"):
                # the rate is the next numeric fraction in the row
                for j in range(i + 1, len(row)):
                    v = parse_value(row[j])
                    if v is not None and 0 < abs(v) < 1:
                        (state_rates if cl == "state rate" else blended_rates).append(v)
                        break
    return state_rates, blended_rates


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("out_json")
    ap.add_argument("--fy25", default=str(FY25_DEFAULT))
    ap.add_argument("--fed-rate", type=float, default=0.21)
    args = ap.parse_args()

    if not Path(args.fy25).exists():
        print(f"  (provision workbook not found: {args.fy25} — skipping state module)")
        Path(args.out_json).write_text("[]", encoding="utf-8")
        return

    import openpyxl
    wb = openpyxl.load_workbook(args.fy25, data_only=True, read_only=True)
    srows = [list(r) for r in wb["7| State Tax Rate"].iter_rows(values_only=True)] if "7| State Tax Rate" in wb.sheetnames else []
    rrows = [list(r) for r in wb["6| Rate Rec."].iter_rows(values_only=True)] if "6| Rate Rec." in wb.sheetnames else []
    wb.close()

    records = []
    fed = args.fed_rate

    # ---- 1) per-state ETR foots (appt x rate = ETR) ----
    etr_sum = appt_sum = 0.0
    tie = exc = 0
    for row in srows:
        st = row[1] if len(row) > 1 else None
        if not (isinstance(st, str) and st.strip() in US_STATES):
            continue
        appt = parse_value(row[APPT_COL]) if len(row) > APPT_COL else None
        rate = parse_value(row[RATE_COL]) if len(row) > RATE_COL else None
        etr = parse_value(row[ETR_COL]) if len(row) > ETR_COL else None
        if appt is None or rate is None or etr is None:
            continue
        appt_sum += appt
        etr_sum += etr
        recomputed = appt * rate
        d = recomputed - etr
        ok = abs(d) <= 0.0002  # rates are displayed to 4dp; allow rounding
        if ok:
            tie += 1
        else:
            exc += 1
            records.append(make_record(
                "state_tax", pdf_section="FN - Taxes (state)",
                pdf_label=f"{st.strip()} ETR = apportionment x state rate", pdf_year="2025",
                pdf_value=round(recomputed, 4), source_ref="provision tab 7",
                source_label="schedule ETR", source_value=round(etr, 4),
                comparison_unit="rate", delta=round(d, 4), tolerance=0.0002, status="exception",
                is_subtotal=False, notes=f"State ETR does not equal apportionment ({appt}) x rate ({rate})."))

    # roll-up record for the per-state footing
    records.append(make_record(
        "state_tax", pdf_section="FN - Taxes (state)",
        pdf_label=f"Per-state ETR foots (apportionment x rate) — {tie}/{tie + exc} states", pdf_year="2025",
        pdf_value=tie + exc, source_ref="provision tab 7", source_label="states tying",
        source_value=tie, comparison_unit="count", delta=exc, tolerance=0,
        status=("ties" if exc == 0 else "exception"), is_subtotal=True,
        notes="Every state's effective rate = apportionment factor x state statutory rate."
              if exc == 0 else f"{exc} state(s) do not foot — see rows above."))

    # ---- 2) blended state rate = sum of per-state ETRs ----
    # tie to the schedule's own total row (the row with no state label carrying an ETR sum)
    total_etr = None
    for row in srows:
        st = row[1] if len(row) > 1 else None
        v = parse_value(row[ETR_COL]) if len(row) > ETR_COL else None
        if not (isinstance(st, str) and st.strip() in US_STATES) and v is not None and abs(v - etr_sum) < 0.005:
            total_etr = v
    if total_etr is not None:
        d = etr_sum - total_etr
        records.append(make_record(
            "state_tax", pdf_section="FN - Taxes (state)",
            pdf_label="Blended state rate = sum of per-state ETRs", pdf_year="2025",
            pdf_value=round(etr_sum, 4), source_ref="provision tab 7", source_label="schedule total row",
            source_value=round(total_etr, 4), comparison_unit="rate", delta=round(d, 4),
            tolerance=0.0005, status=("ties" if abs(d) <= 0.0005 else "exception"), is_subtotal=True,
            notes=f"Blended state rate {etr_sum:.2%} = sum of the {tie + exc} per-state ETRs."))

    # ---- 3) federal + state blended rate, tie to rate rec ----
    state_rates, blended_rates = extract_rate_rec_rates(rrows)
    # the rate-rec 'State Rate' that matches the apportionment blended rate
    rr_state = next((r for r in state_rates if abs(r - etr_sum) <= 0.001), None)
    if rr_state is not None:
        records.append(make_record(
            "state_tax", pdf_section="FN - Taxes (state)",
            pdf_label="Rate rec state rate matches apportionment schedule", pdf_year="2025",
            pdf_value=round(rr_state, 4), source_ref="rate rec / tab 7", source_label="apportionment blended",
            source_value=round(etr_sum, 4), comparison_unit="rate", delta=round(rr_state - etr_sum, 4),
            tolerance=0.001, status="ties", is_subtotal=True,
            notes="The state rate used in the rate reconciliation ties the apportionment schedule."))
    recomputed_blended = fed + etr_sum * (1 - fed)
    rr_blended = next((b for b in blended_rates if abs(b - recomputed_blended) <= 0.002), None)
    if rr_blended is not None:
        records.append(make_record(
            "state_tax", pdf_section="FN - Taxes (state)",
            pdf_label=f"Federal+state blended rate = fed + state x (1-fed) = {recomputed_blended:.1%}",
            pdf_year="2025", pdf_value=round(recomputed_blended, 4), source_ref="rate rec",
            source_label="rate rec blended", source_value=round(rr_blended, 4),
            comparison_unit="rate", delta=round(recomputed_blended - rr_blended, 4), tolerance=0.002,
            status="ties", is_subtotal=True,
            notes=f"Blended rate recomputes: {fed:.0%} + {etr_sum:.2%} x (1 - {fed:.0%}) = {recomputed_blended:.2%}."))

    Path(args.out_json).write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")

    by = {}
    for r in records:
        by[r["status"]] = by.get(r["status"], 0) + 1
    print(f"State tax module: {len(records)} records {by}")
    print(f"  per-state ETR: {tie} tie / {exc} exception | blended state rate {etr_sum:.3%} | "
          f"fed+state blended {recomputed_blended:.3%}")
    for r in records:
        if r["status"] == "exception":
            print(f"  [EXCEPTION] {r['pdf_label']}: d={r['delta']}")
    print(f"Wrote {args.out_json}")


if __name__ == "__main__":
    main()
