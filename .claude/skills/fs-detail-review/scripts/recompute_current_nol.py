"""Provision recompute engine — Module D: current tax + NOL.

Three mechanically-clean, high-value checks (the intricate per-jurisdiction current
split and the NOL vintage roll — which carry return-to-provision true-ups and
apportionment judgment — are deliberately left to the next tier):

  1. TAXABLE-INCOME BUILD FOOTS — pretax book income + permanent differences +
     temporary differences = taxable income, recomputed from the current-tax tab and
     checked against the preparer's taxable-income line. The core current-tax mechanic.

  2. FEDERAL NOL POSITION — when federal taxable income is a loss, federal cash tax
     should be ~0 (the loss is absorbed / becomes an NOL). Confirms the provision isn't
     accruing federal cash tax on a loss year.

  3. NOL DTA TREND — federal + state NOL DTA tracked across FY23/24/25. A
     loss-generating company's NOL should accrete; a *decrease* without taxable income
     (usage / expiry / write-off) is flagged for explanation.

Emits records under lane "current_nol" -> the "Current & NOL" report tab.

CLI:
  python recompute_current_nol.py <out.json>
       [--fy23 <xlsx>] [--fy24 <xlsx>] [--fy25 <xlsx>]
"""

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from tie_out_common import parse_value, make_record  # noqa: E402

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

ROOT = Path(r"C:\path\to\financial-statement-review")
DEFAULTS = {
    "FY23": ROOT / "Prior Year Examples/2023/FS Compilation Requests (FS Compilation Partner)/8. (YE) Acme Corp 2023 Tax Provision_v2024.5.23.xlsx",
    "FY24": ROOT / "PBCs/2024 Audit/FS Compilation Requests (FS Compilation Partner)/Tax Provision/2024 Acme Corp Tax Provision_v2025.5.27.xlsx",
    "FY25": ROOT / "PBCs/2025 Audit/Year End/FS Compilation Requests (FS Compilation Partner) (old)/9. Tax Provision/9. Acme Corp 2025 Tax Provision_v2026.5.21.xlsx",
}

# Federal column index on '5| Current'.
FED_COL = 3


def _rowval(rows, label_substrs, col, label_cols=(0, 1)):
    """Value at `col` on the first row whose label cell (cols 0/1) matches all substrs."""
    for row in rows:
        label = ""
        for lc in label_cols:
            if len(row) > lc and isinstance(row[lc], str) and row[lc].strip():
                label = row[lc]
                break
        ll = label.lower()
        if label and all(s.lower() in ll for s in label_substrs):
            v = parse_value(row[col]) if len(row) > col else None
            if v is not None:
                return v
    return None


def extract_current_build(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows = [list(r) for r in wb["5| Current"].iter_rows(values_only=True)] if "5| Current" in wb.sheetnames else []
    wb.close()
    return {
        "pretax": _rowval(rows, ["pre-tax book income"], FED_COL),
        # the perm/temp TOTAL rows carry a federal-column value (the detail header rows don't)
        "perms": _rowval(rows, ["permanent differences"], FED_COL, label_cols=(1,)),
        "temps": _rowval(rows, ["total temporary differences"], FED_COL),
        "taxable_income": _rowval(rows, ["taxable income"], FED_COL, label_cols=(0,)),
        "net_fed_current": _rowval(rows, ["net federal tax payable"], FED_COL, label_cols=(1,)),
    }


def extract_nol_dta(path):
    """Fed + state NOL DTA (raw $) from '4| DTA DTL Summary' left block."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    fed = state = 0.0
    if "4| DTA DTL Summary" in wb.sheetnames:
        for row in wb["4| DTA DTL Summary"].iter_rows(values_only=True):
            label = row[1] if len(row) > 1 and isinstance(row[1], str) else ""
            val = parse_value(row[3]) if len(row) > 3 else None
            if val is None or not label:
                continue
            ll = label.lower()
            if "fed nol carryforward" in ll and "fin" not in ll:
                fed += val
            elif "state nol carryforward" in ll:
                state += val
    wb.close()
    return fed, state


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("out_json")
    for yr, p in DEFAULTS.items():
        ap.add_argument(f"--{yr.lower()}", default=str(p))
    args = ap.parse_args()

    cur_path = args.fy25
    records = []

    # ---- 1) Taxable-income build foots ----
    if Path(cur_path).exists():
        b = extract_current_build(cur_path)
        if all(b[k] is not None for k in ("pretax", "perms", "temps", "taxable_income")):
            computed = b["pretax"] + b["perms"] + b["temps"]
            d = computed - b["taxable_income"]
            st = "ties" if abs(d) <= 1.0 else "exception"
            records.append(make_record(
                "current_nol", pdf_section="FN - Taxes (current)",
                pdf_label="Taxable income build (pretax + perms + temps)", pdf_year="2025",
                pdf_value=round(computed, 1), source_ref="provision tab 5",
                source_label="preparer taxable income", source_value=round(b["taxable_income"], 1),
                comparison_unit="$1", delta=round(d, 1), tolerance=1.0, status=st, is_subtotal=True,
                notes=("Federal taxable income recomputes: pretax book income + permanent + "
                       "temporary differences = taxable income.") if st == "ties"
                      else "Taxable-income build does NOT foot — investigate."))

            # ---- 2) Federal NOL position ----
            ti = b["taxable_income"]
            net_fed = b["net_fed_current"]
            if net_fed is not None:
                if ti < 0:
                    ok = abs(net_fed) <= 1.0
                    records.append(make_record(
                        "current_nol", pdf_section="FN - Taxes (current)",
                        pdf_label="Federal NOL position (loss year -> no federal cash tax)",
                        pdf_year="2025", pdf_value=round(net_fed, 1), source_ref="provision tab 5",
                        source_label="expected (~0 on a loss)", source_value=0.0,
                        comparison_unit="$1", delta=round(net_fed, 1), tolerance=1.0,
                        status=("ties" if ok else "exception"), is_subtotal=True,
                        notes=("Federal taxable income is a loss; the federal current tax is ~0 — the "
                               "loss is absorbed / carried forward as NOL (consistent).") if ok else
                              "Federal taxable income is a loss but a non-zero federal current tax is "
                              "accrued — investigate."))
        else:
            print("  (could not extract the taxable-income build from tab 5)")

    # ---- 3) NOL DTA trend (fed + state) ----
    nol = {}
    for yr in ("FY23", "FY24", "FY25"):
        p = getattr(args, yr.lower())
        if Path(p).exists():
            nol[yr] = extract_nol_dta(p)
    yrs = list(nol.keys())
    for which, idx in (("Federal NOL DTA", 0), ("State NOL DTA", 1)):
        series = [nol[y][idx] for y in yrs]
        note = " -> ".join(f"{y} {v:,.0f}" for y, v in zip(yrs, series))
        decreased = len(series) >= 2 and series[-1] < series[0] - 1.0
        status = "nol-trend"
        if decreased:
            status = "nol-decrease"
            note = ("NOL DTA decreased vs a prior year — confirm utilization / expiry / write-off "
                    "(a loss company's NOL normally accretes).  " + note)
        records.append(make_record(
            "current_nol", pdf_section="FN - Taxes (NOL trend)",
            pdf_label=which, pdf_year=" / ".join(yrs),
            pdf_value=round(series[-1], 0) if series else None, source_ref="provision tab 4 (3-yr)",
            source_label="prior year", source_value=round(series[0], 0) if series else None,
            comparison_unit="$1", delta=round(series[-1] - series[0], 0) if len(series) >= 2 else None,
            status=status, is_subtotal=False, notes=note))

    Path(args.out_json).write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")

    by = {}
    for r in records:
        by[r["status"]] = by.get(r["status"], 0) + 1
    print(f"Current & NOL module: {len(records)} checks | {by}")
    for r in records:
        tag = r["status"]
        print(f"  [{tag:<12}] {r['pdf_label'][:46]:<46} value={r['pdf_value']:>14,.0f} "
              f"ref={r['source_value'] if r['source_value'] is not None else '-'}")
    print(f"Wrote {args.out_json}")


if __name__ == "__main__":
    main()
